#!/usr/bin/env python3
"""Fetch infrastructure layers from external sources and push to DC Hub.

Runs on the GitHub Actions runner (reliable egress), NOT on Railway —
Railway's network can't reach several infra sources (geo.dot.gov for gas
pipelines times out; the daily server-side ingest fails every run). The
runner fetches the source, builds compact rows, and POSTs them to the
backend's ingest endpoint, which writes to Neon. Railway never fetches.

Env:
  DCHUB_ADMIN_KEY   backend admin key (required)
  DCHUB_ORIGIN      backend origin (default: Railway production)

Usage:
  python tools/infra_fetch.py gas-pipelines        # default cap
  python tools/infra_fetch.py gas-pipelines 30000  # explicit cap
"""
import gzip
import json
import os
import sys
import time
import urllib.parse
import urllib.request

ORIGIN = os.environ.get("DCHUB_ORIGIN", "https://dchub-backend-production.up.railway.app").rstrip("/")
ADMIN = "".join((os.environ.get("DCHUB_ADMIN_KEY") or "").split())

# EIA Natural Gas Interstate + Intrastate Pipelines (national, ~32,892
# polylines). The old geo.dot.gov service died 2026-06 (its backend DB
# refuses connections → 500) and the HIFLD Hp6G80Pky0om7QvQ copy was
# deleted (400) — this FEMA-published EIA org service (FiaPA4ga0iQKduv3)
# is the live replacement. Fields: Operator, TYPEPIPE, Status.
GAS_SVC = ("https://services2.arcgis.com/FiaPA4ga0iQKduv3/arcgis/rest/services/"
           "Natural_Gas_Interstate_and_Intrastate_Pipelines_1/FeatureServer/0/query")

# EIA Electric Power Transmission Lines (national, ~94,619 polylines). Same
# reliable EIA org (FiaPA4ga0iQKduv3). Attributes only — the transmission_lines
# table stores no geometry — so returnGeometry=false (94k full geometries would
# be huge & needless). Fields: ID, TYPE, STATUS, OWNER, VOLTAGE, SUB_1, SUB_2.
# This 94k set supersedes the stale HIFLD 52k snapshot (clean single source).
TRANSMISSION_SVC = ("https://services2.arcgis.com/FiaPA4ga0iQKduv3/arcgis/rest/"
                    "services/US_Electric_Power_Transmission_Lines/FeatureServer/0/query")

# EIA Power Plants in the US (national, ~13,446 points). Same EIA org. Needs
# point geometry (returnGeometry=true, outSR=4326) for lat/lng. Fields:
# Plant_Code, Plant_Name, Utility_Na, sector_nam, City, County, State, Zip,
# PrimSource, Total_MW. Refreshes the same 13,446 plants (no duplication).
POWER_PLANTS_SVC = ("https://services2.arcgis.com/FiaPA4ga0iQKduv3/arcgis/rest/"
                    "services/Power_Plants_in_the_US/FeatureServer/0/query")


def _get(url, timeout=90, tries=5):
    last = None
    for a in range(tries):
        try:
            with urllib.request.urlopen(url, timeout=timeout) as r:
                return r.read()
        except Exception as e:
            last = e
            print(f"    fetch attempt {a+1}/{tries}: {str(e)[:90]}", flush=True)
            time.sleep(min(2 * (a + 1), 12))
    raise last


def _first_point(geom):
    try:
        paths = geom.get("paths") or []
        if paths and paths[0]:
            x, y = paths[0][0][0], paths[0][0][1]
            return float(y), float(x)  # lat, lng
    except Exception:
        pass
    return None, None


def fetch_gas_pipelines(cap):
    """Paginate the geo.dot.gov gas-pipeline service → [[lat,lng,operator,type],...]."""
    rows, offset, page = [], 0, 2000   # service maxRecordCount = 2000
    while len(rows) < cap:
        params = urllib.parse.urlencode({
            "where": "1=1",
            "outFields": "Operator,TYPEPIPE,Status",
            "returnGeometry": "true",
            "outSR": "4326",
            "maxAllowableOffset": "0.05",   # simplify; we keep 1 vertex/line
            "geometryPrecision": "4",
            "resultOffset": offset,
            "resultRecordCount": page,
            "f": "json",
        })
        data = json.loads(_get(GAS_SVC + "?" + params).decode())
        feats = data.get("features") or []
        if not feats:
            break
        for f in feats:
            lat, lng = _first_point(f.get("geometry") or {})
            if lat is None:
                continue
            a = f.get("attributes") or {}
            rows.append([lat, lng, (a.get("Operator") or "")[:200], (a.get("TYPEPIPE") or "")[:80]])
        offset += page
        if len(feats) < page:
            break
    return rows[:cap]


def _v_voltage(v):
    """VOLTAGE is kV int; -999999 / negatives are 'not available' sentinels."""
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _v_clean(s, n):
    if s is None:
        return None
    s = str(s).strip()
    if not s or s.upper() in ("NOT AVAILABLE", "UNKNOWN", "NULL", "NONE"):
        return None
    return s[:n]


# EIA's power-plant service returns the FULL state name ("Mississippi"), but the
# power_plants_eia.state column is varchar(10) and the rest of the schema stores
# 2-letter USPS codes — so a full name >10 chars 500'd the whole batch insert.
# Convert to the postal code (fall back to a 10-char truncation for anything
# unrecognized so the insert can never overflow).
_US_STATE2 = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "district of columbia": "DC", "florida": "FL", "georgia": "GA", "hawaii": "HI",
    "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI",
    "south carolina": "SC", "south dakota": "SD", "tennessee": "TN", "texas": "TX",
    "utah": "UT", "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "puerto rico": "PR", "guam": "GU", "virgin islands": "VI",
}


def _us_state_code(v):
    """Full state name -> 2-letter USPS code (varchar(10)-safe)."""
    s = _v_clean(v, 80)
    if s is None:
        return None
    return _US_STATE2.get(s.lower(), s[:10])


def fetch_transmission_lines(cap):
    """Paginate the EIA transmission service → row tuples for the ingest endpoint.

    Attributes only (returnGeometry=false). Row shape (matches
    routes/transmission_ingest._ROW_FIELDS):
      [hifld_id, name, operator, voltage_kv, from_sub, to_sub, status, line_type]
    """
    rows, offset, page = [], 0, 2000   # service maxRecordCount = 2000
    while len(rows) < cap:
        params = urllib.parse.urlencode({
            "where": "1=1",
            "outFields": "ID,TYPE,STATUS,OWNER,VOLTAGE,SUB_1,SUB_2",
            "returnGeometry": "false",
            "resultOffset": offset,
            "resultRecordCount": page,
            "f": "json",
        })
        data = json.loads(_get(TRANSMISSION_SVC + "?" + params).decode())
        feats = data.get("features") or []
        if not feats:
            break
        for f in feats:
            a = f.get("attributes") or {}
            hid = _v_clean(a.get("ID"), 64)
            owner = _v_clean(a.get("OWNER"), 200)
            name = owner or hid
            rows.append([
                hid,
                name[:200] if name else None,
                owner,
                _v_voltage(a.get("VOLTAGE")),
                _v_clean(a.get("SUB_1"), 200),
                _v_clean(a.get("SUB_2"), 200),
                _v_clean(a.get("STATUS"), 80),
                _v_clean(a.get("TYPE"), 80),
            ])
        offset += page
        if len(feats) < page:
            break
    return rows[:cap]


def fetch_power_plants(cap):
    """Paginate the EIA power-plants service → row tuples for the ingest endpoint.

    Needs point geometry (returnGeometry=true, outSR=4326). Skips non-integer
    Plant_Code. Row shape (matches routes/power_plants_ingest._ROW_FIELDS):
      [plant_id, name, utility_name, state, city, county, zipcode,
       lat, lng, primary_fuel, nameplate_capacity_mw, sector]
    """
    def _vint(v):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None

    def _vfloat(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    rows, offset, page = [], 0, 2000   # service maxRecordCount = 2000
    while len(rows) < cap:
        params = urllib.parse.urlencode({
            "where": "1=1",
            "outFields": ("Plant_Code,Plant_Name,Utility_Na,sector_nam,City,"
                          "County,State,Zip,PrimSource,Total_MW"),
            "returnGeometry": "true",
            "outSR": "4326",
            "resultOffset": offset,
            "resultRecordCount": page,
            "f": "json",
        })
        data = json.loads(_get(POWER_PLANTS_SVC + "?" + params).decode())
        feats = data.get("features") or []
        if not feats:
            break
        for f in feats:
            a = f.get("attributes") or {}
            pid = _vint(a.get("Plant_Code"))
            if pid is None:
                continue
            g = f.get("geometry") or {}
            x, y = g.get("x"), g.get("y")
            lat = _vfloat(y)
            lng = _vfloat(x)
            rows.append([
                pid,
                _v_clean(a.get("Plant_Name"), 200),
                _v_clean(a.get("Utility_Na"), 200),
                _us_state_code(a.get("State")),
                _v_clean(a.get("City"), 100),
                _v_clean(a.get("County"), 100),
                _v_clean(a.get("Zip"), 20),
                lat,
                lng,
                _v_clean(a.get("PrimSource"), 80),
                _vfloat(a.get("Total_MW")),
                _v_clean(a.get("sector_nam"), 120),
            ])
        offset += page
        if len(feats) < page:
            break
    return rows[:cap]


def post_rows(path, rows, cap):
    payload = {"rows": rows}
    body = gzip.compress(json.dumps(payload).encode("utf-8"))
    url = f"{ORIGIN}{path}?cap={cap}"
    req = urllib.request.Request(url, data=body, method="POST", headers={
        "X-Admin-Key": ADMIN, "Content-Type": "application/json",
        "Content-Encoding": "gzip"})
    with urllib.request.urlopen(req, timeout=180) as r:
        return getattr(r, "status", r.getcode()), json.loads(r.read()), len(body)


# EIA-860M operating-generator-capacity (monthly JSON v2 API, not ArcGIS) —
# OPERABLE generator inventory by balancing authority = grid-capacity signal
# ("which grids hold how much of each generation type, incl. standby/reserve").
# EIA's v2 API has NO planned/under-construction route, so this is the operable
# snapshot (the forward queue lives in interconnect_queue). Needs EIA_API_KEY.
# Returns dict rows; the ingest endpoint accepts named fields. Latest period only.
EIA_KEY = os.environ.get("EIA_API_KEY") or os.environ.get("EIA_KEY") or ""
EIA_GEN = "https://api.eia.gov/v2/electricity/operating-generator-capacity/data/"
# OP=operating, SB=standby/reserve, OA=out-of-service but expected back <1yr.
_INVENTORY_STATUS = ["OP", "SB", "OA"]


def fetch_generator_inventory(cap):
    """Latest-period operable generators (operating + standby + returning) from EIA-860M."""
    if not EIA_KEY:
        print("::warning::EIA_API_KEY not set — generator-inventory skipped", flush=True)
        return []
    rows, offset, page, latest = [], 0, 5000, None
    while len(rows) < cap and offset < 80000:
        params = [("api_key", EIA_KEY), ("frequency", "monthly"),
                  ("data[]", "nameplate-capacity-mw"),
                  ("sort[0][column]", "period"), ("sort[0][direction]", "desc"),
                  ("offset", str(offset)), ("length", str(page))]
        for s in _INVENTORY_STATUS:
            params.append(("facets[status][]", s))
        raw = _get(EIA_GEN + "?" + urllib.parse.urlencode(params), timeout=60)
        data = (json.loads(raw).get("response") or {}).get("data") or []
        if not data:
            break
        if latest is None:
            latest = data[0].get("period")   # newest period in the desc-sorted result
        for r in data:
            if r.get("period") != latest:     # sorted desc → past the current snapshot
                return rows[:cap]
            try:
                mw = float(r.get("nameplate-capacity-mw") or 0)
            except (TypeError, ValueError):
                mw = 0.0
            rows.append({
                "period":        (r.get("period") or "")[:7],
                "plant_id":      str(r.get("plantid") or "")[:20],
                "generator_id":  str(r.get("generatorid") or "")[:20],
                "plant_name":    (r.get("plantName") or "")[:200],
                "state":         (r.get("stateid") or "")[:2],
                "ba_code":       (r.get("balancing_authority_code") or "")[:20],
                "entity_name":   (r.get("entityName") or "")[:200],
                "technology":    (r.get("technology") or "")[:100],
                "energy_source": (r.get("energy_source_code") or "")[:20],
                "capacity_mw":   mw,
                "status":        (r.get("status") or "")[:8],
                "status_desc":   (r.get("statusDescription") or "")[:200],
            })
            if len(rows) >= cap:
                break
        offset += page
        if len(data) < page:
            break
    return rows[:cap]


# EIA-860M "Planned" sheet — the forward pipeline of new generators NATIONWIDE
# (incl. non-ISO regions the per-ISO queue misses). The v2 JSON API has no
# planned route, so we parse the monthly Excel here on the runner (13MB +
# openpyxl is too heavy for the 1-replica backend). eia.gov needs a browser UA
# + Referer or the xlsx download returns 0 bytes.
EIA860M_PAGE = "https://www.eia.gov/electricity/data/eia860m/"
_EIA_BUA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"


def _eia_get(url, referer=None, timeout=90, tries=4):
    headers = {"User-Agent": _EIA_BUA}
    if referer:
        headers["Referer"] = referer
    last = None
    for a in range(tries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:
            last = e
            print(f"    eia fetch {a+1}/{tries}: {str(e)[:90]}", flush=True)
            time.sleep(min(2 * (a + 1), 10))
    raise last


def fetch_planned_generators(cap):
    """Download the latest EIA-860M Excel + parse the 'Planned' sheet → dict rows."""
    import io
    import re
    import calendar
    try:
        import openpyxl
    except ImportError:
        print("::error::openpyxl not installed (pip install openpyxl)", flush=True)
        return []
    # Pick the newest non-archive monthly file (e.g. /xls/april_generator2026.xlsx).
    html = _eia_get(EIA860M_PAGE, timeout=40).decode("utf-8", "ignore")
    found = re.findall(r"/electricity/data/eia860m/xls/([a-z]+)_generator(\d{4})\.xlsx", html, re.I)
    months = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
    cand = sorted(set(found), key=lambda x: (int(x[1]), months.get(x[0].lower(), 0)), reverse=True)
    xlsx = None
    for mon, yr in cand:
        url = f"{EIA860M_PAGE}xls/{mon}_generator{yr}.xlsx"
        try:
            raw = _eia_get(url, referer=EIA860M_PAGE, timeout=120)
        except Exception:
            continue
        if raw and len(raw) > 1_000_000:
            print(f"    EIA-860M file: {mon}_generator{yr}.xlsx ({len(raw)//1024//1024}MB)", flush=True)
            xlsx = raw
            break
    if not xlsx:
        print("::error::no usable EIA-860M xlsx found", flush=True)
        return []
    wb = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    if "Planned" not in wb.sheetnames:
        print("::error::no Planned sheet", flush=True)
        return []
    xr = list(wb["Planned"].iter_rows(values_only=True))
    hdr = [str(c) if c is not None else "" for c in xr[2]]   # header on row 3

    def ci(sub):
        for i, h in enumerate(hdr):
            if sub.lower() in h.lower():
                return i
        return None
    c = {k: ci(k) for k in ("Entity Name", "Plant ID", "Plant Name", "Plant State",
                            "County", "Balancing Authority", "Generator ID",
                            "Nameplate Capacity", "Technology", "Energy Source Code",
                            "Status", "Operation Month", "Operation Year",
                            "Latitude", "Longitude")}

    def g(r, key):
        i = c[key]
        return r[i] if i is not None and i < len(r) else None
    out = []
    for r in xr[3:]:
        if not r or not g(r, "Plant ID"):
            continue
        out.append({
            "entity_name":   g(r, "Entity Name"),
            "plant_id":      g(r, "Plant ID"),
            "plant_name":    g(r, "Plant Name"),
            "state":         g(r, "Plant State"),
            "county":        g(r, "County"),
            "ba_code":       g(r, "Balancing Authority"),
            "generator_id":  g(r, "Generator ID"),
            "technology":    g(r, "Technology"),
            "energy_source": g(r, "Energy Source Code"),
            "capacity_mw":   g(r, "Nameplate Capacity"),
            "status":        g(r, "Status"),
            "planned_month": g(r, "Operation Month"),
            "planned_year":  g(r, "Operation Year"),
            "lat":           g(r, "Latitude"),
            "lng":           g(r, "Longitude"),
        })
        if len(out) >= cap:
            break
    return out


LAYERS = {
    "planned-generators": {
        "fetch": fetch_planned_generators,
        "ingest": "/api/v1/admin/ingest/planned-generators",
        "default_cap": 10000,    # EIA-860M Planned sheet is ~2,300 generators
    },
    "generator-inventory": {
        "fetch": fetch_generator_inventory,
        "ingest": "/api/v1/admin/ingest/generator-inventory",
        "default_cap": 40000,    # ~25-30k operable generators nationally
    },
    "gas-pipelines": {
        "fetch": fetch_gas_pipelines,
        "ingest": "/api/v1/admin/ingest/gas-pipelines",
        "default_cap": 30000,
    },
    "transmission-lines": {
        "fetch": fetch_transmission_lines,
        "ingest": "/api/v1/admin/ingest/transmission-lines",
        "default_cap": 100000,   # EIA service has ~94,619 lines
    },
    "power-plants": {
        "fetch": fetch_power_plants,
        "ingest": "/api/v1/admin/ingest/power-plants",
        "default_cap": 20000,    # EIA service has ~13,446 plants
    },
}


# ─────────────────────────────────────────────────────────────────────
# DEAD-MAN BOARD BEAT — carry the REAL inserted count onto /api/v1/ops/deadman.
# The off-worker watcher (tools/deadman/watch.py) beats these same feeds with
# LIVENESS ONLY (last-success ts, no rows_inserted), so every board feed reads
# rows_inserted=NULL and the "0 rows landed" alarm can NEVER trip — #1691's
# COALESCE preserves a real count IF a job posts one, but no board-feed job did.
# This posts the count the ingest endpoint just returned, keyed to the SAME feed
# name the board/watcher use (the GitHub workflow filename, .yml stripped), so a
# "green" feed finally means "rows actually landed", not just "workflow exited 0".
# Fail-OPEN: a beat error is swallowed and NEVER changes this job's exit code. We
# deliberately DON'T send cadence_hours — tools/deadman/watch.py is the
# authoritative cadence registry, and the beat endpoint COALESCEs a missing
# cadence, so its value is left untouched.
# ─────────────────────────────────────────────────────────────────────

# layer key (this script) -> board feed name. The board feed name is the GitHub
# workflow filename that runs `infra_fetch.py <layer>`, .yml stripped — i.e. the
# exact key tools/deadman/watch.py registers. Only layers whose board workflow
# actually invokes THIS script are mapped; any other layer skips the beat (no
# double-reporting, fail-open).
_BOARD_FEED = {
    "gas-pipelines":      "gas-pipeline-ingest",       # gas-pipeline-ingest.yml
    "transmission-lines": "transmission-ingest",       # transmission-ingest.yml
    "power-plants":       "power-plants-ingest",        # power-plants-ingest.yml
    "planned-generators": "planned-generators-ingest",  # planned-generators-ingest.yml
}


def beat_board(layer, rows_inserted, status="success"):
    """Best-effort dead-man BEAT carrying the REAL inserted count. NEVER raises."""
    feed = _BOARD_FEED.get(layer)
    if not feed or not ADMIN:
        return
    try:
        body = json.dumps({
            "feed": feed,
            "status": status,
            "rows_inserted": int(rows_inserted or 0),
            "note": "real inserted count from infra_fetch.py (%s)" % layer,
        }).encode()
        # Runs on the GitHub Actions runner, NOT the worker — beat the backend
        # ORIGIN directly (same origin + admin key the ingest POST already uses).
        req = urllib.request.Request(
            f"{ORIGIN}/api/v1/admin/ingest-runs/beat", data=body, method="POST",
            headers={"Content-Type": "application/json",
                     "User-Agent": "dchub-infra-fetch/1.0",
                     "X-Admin-Key": ADMIN})
        urllib.request.urlopen(req, timeout=20).read()
        print(f"  board beat: {feed} rows_inserted={int(rows_inserted or 0)}", flush=True)
    except Exception as e:  # fail-open: a beat error must never fail the ingest
        print(f"  board beat failed (non-fatal): {str(e)[:120]}", flush=True)


def main():
    if not ADMIN:
        print("::error::missing DCHUB_ADMIN_KEY"); return 1
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    layer = args[0] if args else "gas-pipelines"
    if layer not in LAYERS:
        print(f"::error::unknown layer {layer}; known: {list(LAYERS)}"); return 1
    cfg = LAYERS[layer]
    cap = int(args[1]) if len(args) > 1 else cfg["default_cap"]

    t0 = time.time()
    print(f"== {layer}: fetching (cap {cap}) ==", flush=True)
    rows = cfg["fetch"](cap)
    print(f"  fetched {len(rows):,} points in {time.time()-t0:.0f}s", flush=True)
    if not rows:
        print("::error::source returned 0 rows"); return 1
    st, j, gz = post_rows(cfg["ingest"], rows, cap)
    print(f"  posted {gz//1024}KB gz (HTTP {st}) → {json.dumps(j)[:200]}", flush=True)
    if st != 200 or not j.get("ok"):
        print(f"::error::ingest failed (HTTP {st})"); return 1
    print(f"== {layer} done: {j.get('inserted')} rows inserted ==")
    # Additive, fail-open: publish the REAL inserted count to the dead-man board.
    beat_board(layer, j.get("inserted"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
