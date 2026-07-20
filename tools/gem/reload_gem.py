"""GEM re-ingest — download the bundle from R2 and load all five datasets into Neon.

The runner CAN reach R2 (unlike GEM's gated site). Owner drops a fresh gem-data.zip
at r2://<GEM_R2_BUCKET>/<GEM_R2_KEY> each release; this loads it. Glob-based file
discovery so it survives GEM's month-in-filename changes. Every load is a full-replace
by source tag → idempotent. Fails (exit 1) if the core power file is missing/empty so
the workflow goes red instead of silently wiping.

Env: DATABASE_URL (or NEON_DATABASE_URL), R2_ENDPOINT_URL, R2_ACCESS_KEY_ID,
R2_SECRET_ACCESS_KEY, GEM_R2_BUCKET (default dchub-backups), GEM_R2_KEY (default gem/gem-data.zip).
"""
import os, sys, json, glob, zipfile, tempfile
import openpyxl, psycopg2, boto3
from botocore.config import Config
from psycopg2.extras import execute_values

DSN = os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL")
BUCKET = os.environ.get("GEM_R2_BUCKET", "dchub-backups")
KEY = os.environ.get("GEM_R2_KEY", "gem/gem-data.zip")


def num(v):
    try:
        f = float(v); return f if f == f else None
    except (TypeError, ValueError):
        return None


def s(v, cap):
    return "" if v is None else str(v)[:cap]


def coords_of(geom):
    if not geom:
        return
    if geom.get("type") == "GeometryCollection":
        for g in geom.get("geometries", []):
            yield from coords_of(g)
        return
    def walk(c):
        if not c:
            return
        if isinstance(c[0], (int, float)):
            if len(c) >= 2:
                yield (c[0], c[1])
        else:
            for x in c:
                yield from walk(x)
    yield from walk(geom.get("coordinates"))


def conn():
    return psycopg2.connect(DSN, sslmode="require", connect_timeout=20)


# ── R2 + extraction ─────────────────────────────────────────────────────────
def fetch_bundle(wd):
    s3 = boto3.client("s3", endpoint_url=os.environ["R2_ENDPOINT_URL"],
                      aws_access_key_id=os.environ["R2_ACCESS_KEY_ID"],
                      aws_secret_access_key=os.environ["R2_SECRET_ACCESS_KEY"],
                      config=Config(signature_version="s3v4"))
    zp = os.path.join(wd, "gem-data.zip")
    print(f"downloading r2://{BUCKET}/{KEY} …", flush=True)
    s3.download_file(BUCKET, KEY, zp)
    with zipfile.ZipFile(zp) as z:
        z.extractall(wd)
    return wd


def find(wd, pattern):
    m = [x for x in glob.glob(os.path.join(wd, "**", pattern), recursive=True) if "__MACOSX" not in x]
    return m[0] if m else None


def nested_geojson(wd, zip_pattern, name_contains=None):
    """Find a nested .zip, extract it, return the .geojson inside (optionally name-filtered)."""
    nz = find(wd, zip_pattern)
    if not nz:
        return None
    out = nz + "_x"; os.makedirs(out, exist_ok=True)
    with zipfile.ZipFile(nz) as z:
        z.extractall(out)
    cands = [x for x in glob.glob(os.path.join(out, "**", "*.geojson"), recursive=True) if "__MACOSX" not in x]
    if name_contains:
        cands = [c for c in cands if name_contains.lower() in os.path.basename(c).lower()]
    cands.sort(key=lambda p: os.path.getsize(p), reverse=True)
    return cands[0] if cands else None


# ── loaders (idempotent full-replace by source) ─────────────────────────────
def _batch(cur, table, cols, rows):
    ph = "(" + ",".join(["%s"] * len(cols)) + ")"
    execute_values(cur, f"INSERT INTO {table} ({','.join(cols)}) VALUES %s",
                   rows, template=None, page_size=500)


def load_power(xlsx):
    SRC = "gem_integrated_power"
    CM = {"GEM unit/phase ID": "gem_id", "Type": "fuel_type", "Plant / Project name": "plant_name",
          "Unit / Phase name": "unit_name", "Capacity (MW)": "capacity_mw", "Status": "status",
          "Start year": "start_year", "Technology": "technology", "Country/area": "country",
          "Region": "region", "Operator(s)": "operator", "Owner(s)": "owner",
          "Latitude": "lat", "Longitude": "lng", "GEM.Wiki URL": "wiki_url"}
    F = ["gem_id", "fuel_type", "plant_name", "unit_name", "capacity_mw", "status", "start_year",
         "technology", "country", "region", "operator", "owner", "lat", "lng", "wiki_url"]
    NUMS = {"capacity_mw", "start_year", "lat", "lng"}
    CAPS = {"gem_id": 40, "fuel_type": 40, "plant_name": 250, "unit_name": 150, "status": 60,
            "technology": 120, "country": 100, "region": 80, "operator": 200, "owner": 200, "wiki_url": 250}
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Power facilities"]; it = ws.iter_rows(values_only=True)
    hdr = list(next(it)); ix = {h: i for i, h in enumerate(hdr)}
    rows = []
    for r in it:
        lat, lng = num(r[ix["Latitude"]]), num(r[ix["Longitude"]])
        if lat is None or lng is None:
            continue
        rec = {}
        for h, c in CM.items():
            v = r[ix[h]]
            rec[c] = num(v) if c in NUMS else s(v, CAPS.get(c, 200))
        rec["lat"], rec["lng"] = lat, lng
        rows.append(tuple(rec.get(f) for f in F) + (SRC,))
    wb.close()
    with conn() as c, c.cursor() as cur:
        cur.execute("""CREATE TABLE IF NOT EXISTS gem_power (id SERIAL PRIMARY KEY, gem_id TEXT, fuel_type TEXT,
          plant_name TEXT, unit_name TEXT, capacity_mw NUMERIC, status TEXT, start_year NUMERIC, technology TEXT,
          country TEXT, region TEXT, operator TEXT, owner TEXT, lat DOUBLE PRECISION, lng DOUBLE PRECISION,
          wiki_url TEXT, source TEXT, ingested_at TIMESTAMPTZ DEFAULT NOW())""")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_gempow_bbox ON gem_power(lng,lat)")
        cur.execute("DELETE FROM gem_power WHERE source=%s", (SRC,))
        _batch(cur, "gem_power", F + ["source"], rows)
        c.commit()
    return len(rows)


def load_lng(xlsx):
    SRC = "gem_gas_infra"
    F = ["gem_id", "kind", "name", "unit_name", "fuel", "capacity", "capacity_units", "status",
         "start_year", "country", "region", "owner", "lat", "lng", "wiki_url"]
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["LNG Terminals"]; it = ws.iter_rows(values_only=True)
    hdr = list(next(it)); ix = {h: i for i, h in enumerate(hdr)}
    g = lambda r, h: r[ix[h]] if h in ix else None
    rows = []
    for r in it:
        lat, lng = num(g(r, "Latitude")), num(g(r, "Longitude"))
        if lat is None or lng is None:
            continue
        rows.append((s(g(r, "UnitID") or g(r, "ProjectID"), 40), "lng_terminal", s(g(r, "TerminalName"), 250),
                     s(g(r, "UnitName") or g(r, "FacilityType"), 150), s(g(r, "Fuel") or "LNG", 40),
                     num(g(r, "CapacityinMtpa")) or num(g(r, "Capacity")),
                     "Mtpa" if num(g(r, "CapacityinMtpa")) else s(g(r, "CapacityUnits"), 30),
                     s(g(r, "Status"), 60), num(g(r, "ActualStartYear")) or num(g(r, "LatestPlannedStartYear")),
                     s(g(r, "Country/Area"), 100), s(g(r, "Region"), 80), s(g(r, "Owner"), 200),
                     lat, lng, s(g(r, "Wiki"), 250), SRC))
    wb.close()
    with conn() as c, c.cursor() as cur:
        cur.execute("""CREATE TABLE IF NOT EXISTS gem_gas (id SERIAL PRIMARY KEY, gem_id TEXT, kind TEXT, name TEXT,
          unit_name TEXT, fuel TEXT, capacity NUMERIC, capacity_units TEXT, status TEXT, start_year NUMERIC,
          country TEXT, region TEXT, owner TEXT, lat DOUBLE PRECISION, lng DOUBLE PRECISION, wiki_url TEXT,
          source TEXT, ingested_at TIMESTAMPTZ DEFAULT NOW())""")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_gemgas_bbox ON gem_gas(lng,lat)")
        cur.execute("DELETE FROM gem_gas WHERE source=%s", (SRC,))
        _batch(cur, "gem_gas", F + ["source"], rows)
        c.commit()
    return len(rows)


def load_pipelines(geojson, src):
    d = json.load(open(geojson)); rows = []
    for f in d.get("features", []):
        g = f.get("geometry"); p = f.get("properties", {}) or {}
        pts = list(coords_of(g)) if g else []
        if not pts:
            continue
        lngs = [c[0] for c in pts]; lats = [c[1] for c in pts]
        rows.append((s(p.get("ProjectID"), 40), s(p.get("PipelineName"), 250), s(p.get("SegmentName"), 200),
                     s(p.get("Status"), 60), s(p.get("Fuel"), 40), s(p.get("CountriesOrAreas"), 200),
                     s(p.get("Owner"), 200), num(p.get("StartYear1")), json.dumps(g, separators=(",", ":")),
                     min(lngs), min(lats), max(lngs), max(lats), src))
    cols = ["project_id", "name", "segment", "status", "fuel", "countries", "owner", "start_year",
            "geom_json", "min_lng", "min_lat", "max_lng", "max_lat", "source"]
    with conn() as c, c.cursor() as cur:
        cur.execute("""CREATE TABLE IF NOT EXISTS gem_gas_pipelines (id SERIAL PRIMARY KEY, project_id TEXT, name TEXT,
          segment TEXT, status TEXT, fuel TEXT, countries TEXT, owner TEXT, start_year NUMERIC, geom_json TEXT,
          min_lng DOUBLE PRECISION, min_lat DOUBLE PRECISION, max_lng DOUBLE PRECISION, max_lat DOUBLE PRECISION,
          source TEXT, ingested_at TIMESTAMPTZ DEFAULT NOW())""")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_gempipe_bbox ON gem_gas_pipelines(min_lng,max_lng,min_lat,max_lat)")
        cur.execute("DELETE FROM gem_gas_pipelines WHERE source=%s", (src,))
        _batch(cur, "gem_gas_pipelines", cols, rows)
        c.commit()
    return len(rows)


def load_coalmines(geojson):
    SRC = "gem_coal_mines"; d = json.load(open(geojson)); rows = []
    for f in d.get("features", []):
        g = f.get("geometry"); p = f.get("properties", {}) or {}
        pts = list(coords_of(g)) if g else []
        if not pts:
            continue
        lngs = [c[0] for c in pts]; lats = [c[1] for c in pts]
        rows.append((s(p.get("GEM Mine ID"), 40), s(p.get("Mine Name"), 200), s(p.get("mine feature category"), 60),
                     s(p.get("mine feature subcategory"), 80), s(p.get("Coal Grade"), 60), s(p.get("Owners"), 250),
                     s(p.get("Parent Company"), 200), s(p.get("Country / Area"), 100),
                     s(p.get("GEM Wiki Page (ENG)"), 250), json.dumps(g, separators=(",", ":")),
                     min(lngs), min(lats), max(lngs), max(lats), SRC))
    cols = ["gem_mine_id", "mine_name", "category", "subcategory", "coal_grade", "owners", "parent",
            "country", "wiki", "geom_json", "min_lng", "min_lat", "max_lng", "max_lat", "source"]
    with conn() as c, c.cursor() as cur:
        cur.execute("""CREATE TABLE IF NOT EXISTS gem_coal_mines (id SERIAL PRIMARY KEY, gem_mine_id TEXT, mine_name TEXT,
          category TEXT, subcategory TEXT, coal_grade TEXT, owners TEXT, parent TEXT, country TEXT, wiki TEXT,
          geom_json TEXT, min_lng DOUBLE PRECISION, min_lat DOUBLE PRECISION, max_lng DOUBLE PRECISION,
          max_lat DOUBLE PRECISION, source TEXT, ingested_at TIMESTAMPTZ DEFAULT NOW())""")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_gemcm_bbox ON gem_coal_mines(min_lng,max_lng,min_lat,max_lat)")
        cur.execute("DELETE FROM gem_coal_mines WHERE source=%s", (SRC,))
        _batch(cur, "gem_coal_mines", cols, rows)
        c.commit()
    return len(rows)


def main():
    if not DSN:
        print("no DATABASE_URL", file=sys.stderr); sys.exit(2)
    with tempfile.TemporaryDirectory() as wd:
        fetch_bundle(wd)
        res = {}
        f = find(wd, "*Integrated-Power*.xlsx"); res["gem_power"] = load_power(f) if f else "MISSING"
        f = find(wd, "*LNG*T*minals*.xlsx");     res["gem_lng"] = load_lng(f) if f else "MISSING"
        f = nested_geojson(wd, "*Gas-Pipelines*.zip"); res["gas_pipelines"] = load_pipelines(f, "gem_ggit_pipelines") if f else "MISSING"
        f = nested_geojson(wd, "*Oil-NGL*Pipelines*.zip"); res["oil_pipelines"] = load_pipelines(f, "gem_goit_oil_ngl") if f else "MISSING"
        f = nested_geojson(wd, "*Coal Mine Boundaries*.zip", name_contains="Methane"); res["coal_mines"] = load_coalmines(f) if f else "MISSING"
        print("RELOAD RESULT: " + json.dumps(res))
        if not isinstance(res["gem_power"], int) or res["gem_power"] < 1000:
            print("FAIL: core power load missing/short", file=sys.stderr); sys.exit(1)


if __name__ == "__main__":
    main()
