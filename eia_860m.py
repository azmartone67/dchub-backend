"""
EIA-860M Monthly Generator Report - Lat/Lng Extraction
=======================================================
Downloads the latest EIA-860M monthly CSV and extracts exact coordinates
for all utility-scale power plants in the US.

Source: https://www.eia.gov/electricity/data/eia860m/
File:   EIA860m monthly generator inventory (ZIP → CSV)

Key fields:
  - Plant ID, Plant Name, Plant State
  - Latitude, Longitude (generator-level)
  - Nameplate Capacity (MW), Technology, Energy Source
  - Status (OP=Operating, SB=Standby, OA=Out of service, etc.)
"""

import csv
import io
import logging
import os
import tempfile
import zipfile
from datetime import datetime
from typing import Optional

import requests

logger = logging.getLogger(__name__)

# EIA publishes the 860M at a predictable URL pattern
# The file is typically named like: december_generator2024.xlsx or .zip
EIA_860M_BASE_URL = "https://www.eia.gov/electricity/data/eia860m/xls"

# Fallback: direct bulk download page
EIA_BULK_URL = "https://api.eia.gov/bulk/EIA860M.zip"

# Column mappings (EIA changes headers slightly between months)
COLUMN_ALIASES = {
    "plant_id": ["Plant ID", "Plant Id", "Entity ID", "Utility ID", "plant_id"],
    "plant_name": ["Plant Name", "plant_name", "Station Name"],
    "state": ["Plant State", "State", "plant_state"],
    "latitude": ["Latitude", "latitude", "Lat"],
    "longitude": ["Longitude", "longitude", "Lon", "Long"],
    "capacity_mw": ["Nameplate Capacity (MW)", "Net Summer Capacity (MW)",
                     "nameplate_capacity_mw", "Capacity (MW)"],
    "technology": ["Technology", "Prime Mover", "technology"],
    "energy_source": ["Energy Source 1", "Energy Source Code",
                       "energy_source_1", "Fuel Type"],
    "status": ["Status", "Operating Status", "status", "Plant Status"],
    "county": ["County", "county"],
    "balancing_authority": ["Balancing Authority Code", "BA Code",
                             "balancing_authority_code"],
}


class EIA860MIngester:
    """Downloads and parses EIA-860M monthly data for power plant coordinates."""

    def __init__(self, cache_dir: Optional[str] = None, api_key: Optional[str] = None):
        """
        Args:
            cache_dir: Directory to cache downloaded files. Defaults to temp dir.
            api_key: EIA API key (optional, enables bulk API access).
        """
        self.cache_dir = cache_dir or tempfile.mkdtemp(prefix="eia860m_")
        self.api_key = api_key or os.environ.get("EIA_API_KEY")
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "DCHub-DataEnrichment/1.0 (dchub.cloud)"
        })

    def _resolve_column(self, headers: list[str], field_name: str) -> Optional[int]:
        """Find column index by checking known aliases."""
        aliases = COLUMN_ALIASES.get(field_name, [field_name])
        for alias in aliases:
            # Case-insensitive, strip whitespace
            for i, h in enumerate(headers):
                if h.strip().lower() == alias.strip().lower():
                    return i
        return None

    def _build_column_map(self, headers: list[str]) -> dict[str, Optional[int]]:
        """Build a mapping of field_name -> column_index."""
        col_map = {}
        for field in COLUMN_ALIASES:
            idx = self._resolve_column(headers, field)
            col_map[field] = idx
            if idx is None:
                logger.warning(f"Column not found for '{field}' in headers")
        return col_map

    def download_latest_csv(self, year: Optional[int] = None,
                            month: Optional[int] = None) -> str:
        """
        Download the latest EIA-860M ZIP and extract the generator CSV.

        Args:
            year: Target year (defaults to current year).
            month: Target month (defaults to 2 months ago, typical publication lag).

        Returns:
            Path to extracted CSV file.
        """
        now = datetime.now()
        year = year or now.year
        # EIA typically lags ~2 months
        month = month or max(1, now.month - 2)

        month_names = [
            "", "january", "february", "march", "april", "may", "june",
            "july", "august", "september", "october", "november", "december"
        ]
        month_name = month_names[month]

        # Try common URL patterns
        url_patterns = [
            f"{EIA_860M_BASE_URL}/{month_name}_generator{year}.zip",
            f"{EIA_860M_BASE_URL}/{month_name}_generator{year}.xlsx",
            f"{EIA_860M_BASE_URL}/eia860m_{year}_{month:02d}.zip",
        ]

        zip_path = os.path.join(self.cache_dir, f"eia860m_{year}_{month:02d}.zip")

        for url in url_patterns:
            try:
                logger.info(f"Trying EIA-860M download: {url}")
                resp = self.session.get(url, timeout=60)
                if resp.status_code == 200:
                    with open(zip_path, "wb") as f:
                        f.write(resp.content)
                    logger.info(f"Downloaded EIA-860M from {url}")
                    break
            except requests.RequestException as e:
                logger.debug(f"URL failed: {url} -> {e}")
                continue
        else:
            # Fallback: try EIA bulk API
            if self.api_key:
                logger.info("Trying EIA bulk API fallback...")
                bulk_url = f"https://api.eia.gov/v2/electricity/facility-fuel/data/?api_key={self.api_key}&frequency=monthly"
                # Note: Bulk API returns JSON, different parsing needed
                raise NotImplementedError(
                    "Bulk API JSON parsing not yet implemented. "
                    "Provide year/month that has a published ZIP."
                )
            raise FileNotFoundError(
                f"Could not download EIA-860M for {month_name} {year}. "
                f"Check https://www.eia.gov/electricity/data/eia860m/ for available dates."
            )

        # Extract CSV from ZIP
        csv_path = self._extract_csv_from_zip(zip_path)
        return csv_path

    def _extract_csv_from_zip(self, zip_path: str) -> str:
        """Extract the generator CSV/XLSX from a ZIP file."""
        with zipfile.ZipFile(zip_path, "r") as zf:
            # Look for generator-related files
            candidates = [
                n for n in zf.namelist()
                if "generator" in n.lower() and (
                    n.endswith(".csv") or n.endswith(".xlsx")
                )
            ]
            if not candidates:
                # Fall back to any CSV
                candidates = [n for n in zf.namelist() if n.endswith(".csv")]

            if not candidates:
                raise FileNotFoundError(
                    f"No CSV/XLSX found in ZIP. Contents: {zf.namelist()}"
                )

            target = candidates[0]
            extract_path = zf.extract(target, self.cache_dir)
            logger.info(f"Extracted: {target}")

            # If XLSX, convert to CSV using openpyxl
            if target.endswith(".xlsx"):
                extract_path = self._xlsx_to_csv(extract_path)

            return extract_path

    def _xlsx_to_csv(self, xlsx_path: str) -> str:
        """Convert XLSX to CSV for uniform parsing."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("pip install openpyxl --break-system-packages")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        # Find the operating generators sheet
        target_sheet = None
        for name in wb.sheetnames:
            if "operating" in name.lower() or "generator" in name.lower():
                target_sheet = name
                break
        target_sheet = target_sheet or wb.sheetnames[0]

        ws = wb[target_sheet]
        csv_path = xlsx_path.replace(".xlsx", ".csv")

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        wb.close()
        logger.info(f"Converted XLSX → CSV: {csv_path}")
        return csv_path

    def parse_plants(self, csv_path: str) -> list[dict]:
        """
        Parse the EIA-860M CSV and return plant records with coordinates.

        Returns:
            List of dicts with standardized plant data including lat/lng.
        """
        plants = {}  # Deduplicate by plant_id (multiple generators per plant)

        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)

            # Find header row (might not be row 0 due to metadata rows)
            headers = None
            col_map = None
            for row in reader:
                if any("plant" in str(cell).lower() for cell in row[:5]):
                    headers = [str(c).strip() for c in row]
                    col_map = self._build_column_map(headers)
                    if col_map.get("latitude") is not None:
                        break

            if not headers or not col_map:
                raise ValueError("Could not find header row in CSV")

            lat_idx = col_map["latitude"]
            lng_idx = col_map["longitude"]
            id_idx = col_map["plant_id"]

            if lat_idx is None or lng_idx is None or id_idx is None:
                raise ValueError(
                    f"Missing critical columns. Found: {col_map}"
                )

            # Parse data rows
            for row in reader:
                if len(row) <= max(lat_idx, lng_idx, id_idx):
                    continue

                try:
                    plant_id = str(row[id_idx]).strip()
                    lat = row[lat_idx].strip() if lat_idx < len(row) else ""
                    lng = row[lng_idx].strip() if lng_idx < len(row) else ""

                    if not plant_id or not lat or not lng:
                        continue

                    lat_f = float(lat)
                    lng_f = float(lng)

                    # Sanity check coordinates (continental US + territories)
                    if not (-90 <= lat_f <= 90 and -180 <= lng_f <= 180):
                        logger.warning(f"Invalid coords for plant {plant_id}: {lat_f}, {lng_f}")
                        continue

                    # Build plant record (take first generator's data per plant)
                    if plant_id not in plants:
                        def _get(field):
                            idx = col_map.get(field)
                            if idx is not None and idx < len(row):
                                return str(row[idx]).strip()
                            return None

                        plants[plant_id] = {
                            "eia_plant_id": plant_id,
                            "name": _get("plant_name"),
                            "state": _get("state"),
                            "county": _get("county"),
                            "latitude": lat_f,
                            "longitude": lng_f,
                            "capacity_mw": _safe_float(_get("capacity_mw")),
                            "technology": _get("technology"),
                            "energy_source": _get("energy_source"),
                            "status": _get("status"),
                            "balancing_authority": _get("balancing_authority"),
                            "source": "eia_860m",
                            "updated_at": datetime.utcnow().isoformat(),
                        }
                    else:
                        # Aggregate capacity across generators at same plant
                        cap = _safe_float(
                            str(row[col_map["capacity_mw"]]).strip()
                            if col_map.get("capacity_mw") is not None
                            and col_map["capacity_mw"] < len(row)
                            else None
                        )
                        if cap and plants[plant_id]["capacity_mw"]:
                            plants[plant_id]["capacity_mw"] += cap

                except (ValueError, IndexError) as e:
                    continue  # Skip malformed rows silently

        result = list(plants.values())
        logger.info(f"Parsed {len(result)} unique plants from EIA-860M")
        return result

    def ingest(self, year: Optional[int] = None,
               month: Optional[int] = None) -> list[dict]:
        """
        Full pipeline: download → parse → return plant records.

        Returns:
            List of plant dicts with lat/lng coordinates.
        """
        csv_path = self.download_latest_csv(year=year, month=month)
        return self.parse_plants(csv_path)


def _safe_float(val) -> Optional[float]:
    """Safely convert to float, return None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
