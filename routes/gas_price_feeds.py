"""
gas_price_feeds.py (2026-07-01)

Two new gas data feeds, following the iso_queue_ingest.py architecture
(each ingestor returns (ok, parsed_dict, debug); INGESTORS registry;
idempotent upserts; cron-fired daily via cron_heartbeat):

  1. HENRY HUB DAILY SPOT — EIA Open Data API v2, route
     natural-gas/pri/fut/data, series RNGWHHD ("Henry Hub Natural Gas
     Spot Price", $/MMBtu). NOTE: despite the "fut" route name this IS
     the daily SPOT series — EIA v2 has no natural-gas/pri/spt route
     (verified live 2026-07-01: /pri children are [sum, fut, rescom] and
     pri/spt returns 400 "Requested path is not valid"). Henry Hub is the
     only FREE daily gas price; regional basis hubs (Waha, DomSouth,
     Chicago, SoCal, Algonquin) are paid-only (NGI/Argus/Platts) and are
     deliberately NOT fetched here.
     Requires EIA_API_KEY (read from env at call time, mirroring
     eia_api.py / eia_gas_prices_loader.py). No key -> ingest fails
     gracefully (no synthetic rows are EVER written).

  2. LNG EXPORT TERMINALS — EIA "U.S. Liquefaction Capacity" workbook
     (quarterly XLSX, verified live 2026-07-01):
       https://www.eia.gov/naturalgas/importsexports/liquefactioncapacity/U.S.liquefactioncapacity.xlsx
     Sheets parsed: "Existing & Under Construction" (per-train rows,
     grouped to per-project) + "Approved" (per-project). as_of = the
     workbook's own "Release Date" (Contents sheet) — honest even when
     EIA lags. This table is the future live source for the map's
     currently hardcoded 10-terminal LNG_TERMINALS layer.

Tables (DDL runs INSIDE the ingest handlers, never at import/boot — the
boot-DDL statement-timeout storm has failed deploys before):
  henry_hub_spot        (period DATE PK, price_usd_mmbtu, source_url, ingested_at)
  lng_export_terminals  (project_name, category, operator, state, trains,
                         baseload_bcfd/mtpa, peak_bcfd/mtpa, status,
                         in_service, as_of, source_url, ingested_at)

Wiring (main.py):
    from routes.gas_price_feeds import gas_feeds_bp
    app.register_blueprint(gas_feeds_bp)

Endpoints:
  POST /api/v1/gas/feeds/ingest         — run both ingestors (cron target)
  POST /api/v1/gas/feeds/ingest/<feed>  — run one (henry_hub | lng_terminals)
  GET  /api/v1/gas/feeds/status         — last-ingest summary
  GET  /api/v1/gas/henry-hub            — latest spot + 1d/7d/30d trend + history
  GET  /api/v1/gas/lng-terminals        — terminals (name/operator/state/capacity/status)

Cron: fired by cron_heartbeat daily at 07:xx UTC (wide window — the
heartbeat is throttled to ~hourly; both ingestors are idempotent).
"""
import os
import re
import io
import json
import datetime
import urllib.request
import urllib.error
import urllib.parse
from flask import Blueprint, jsonify, request
from internal_auth import is_valid_internal_key
import psycopg

gas_feeds_bp = Blueprint("gas_price_feeds", __name__, url_prefix="/api/v1/gas")

NEON_URL = os.environ.get("NEON_DATABASE_URL") or os.environ.get("DATABASE_URL")

UA = "DCHub-GasFeedsIngest/1.0 (+https://dchub.cloud)"

EIA_BASE = "https://api.eia.gov/v2"
HH_SERIES = "RNGWHHD"
HH_ROUTE = "natural-gas/pri/fut/data"   # the daily SPOT series lives here (no /pri/spt in v2)
HH_SOURCE_LABEL = ("EIA Open Data v2 — natural-gas/pri/fut, series RNGWHHD "
                   "(Henry Hub Natural Gas Spot Price, $/MMBtu)")
HH_PUBLIC_URL = "https://www.eia.gov/dnav/ng/hist/rngwhhdD.htm"

LNG_XLSX_URL = ("https://www.eia.gov/naturalgas/importsexports/"
                "liquefactioncapacity/U.S.liquefactioncapacity.xlsx")
LNG_SOURCE_LABEL = "EIA U.S. Liquefaction Capacity (quarterly workbook)"

# DDL gate parity with eia_gas_prices_loader / db_utils (SKIP_DDL=1 -> assume tables exist).
SKIP_DDL = os.environ.get("SKIP_DDL", "0") == "1"

FEED_STATUS = {
    "henry_hub":     "real (EIA v2 natural-gas/pri/fut RNGWHHD daily spot; needs EIA_API_KEY)",
    "lng_terminals": "real (EIA U.S. Liquefaction Capacity XLSX; Existing+UC per-train grouped + Approved)",
}


# ──────────────────────────────────────────────────────────────────────
# helpers
# ──────────────────────────────────────────────────────────────────────
def _fetch(url, timeout=60, return_bytes=False, headers=None):
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.9",
        **(headers or {}),
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read()
        return (body if return_bytes else body.decode("utf-8", errors="replace")), resp.status


def _try_openpyxl(xlsx_bytes):
    try:
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except ImportError:
        return None
    except Exception:
        return None


def _num(v):
    """Parse EIA workbook numerics: floats, or strings like '10.00^' /
    '2.00^' (footnoted Mtpa) — 'Note A' / '-' / '' -> None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    s = s.rstrip("^+*")
    try:
        return float(s)
    except ValueError:
        return None


def _clean_status(s):
    """Strip single-letter footnote suffixes: 'Commercial operationD' ->
    'Commercial operation', 'Under constructionH' -> 'Under construction'."""
    s = str(s or "").strip()
    return re.sub(r"(?<=[a-z])[A-Z]$", "", s)


def _train_count(label):
    """'Train 1' -> 1; 'Trains 1-9C' -> 9; 'Trains 10-18C' -> 9; else 1/None."""
    if not label:
        return None
    m = re.search(r"(\d+)\s*-\s*(\d+)", str(label))
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return max(b - a + 1, 1)
    return 1 if re.search(r"\d", str(label)) else None


def _cell_str(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    s = str(v).strip()
    return s or None


def _ensure_schema(cur):
    """CREATE TABLE IF NOT EXISTS for both feed tables. Called from the
    ingest request handlers only (NEVER at import/boot — boot DDL has
    caused statement-timeout storms that failed deploys)."""
    if SKIP_DDL:
        return
    cur.execute("""
        CREATE TABLE IF NOT EXISTS henry_hub_spot (
            period          DATE PRIMARY KEY,
            price_usd_mmbtu NUMERIC(8,3) NOT NULL,
            source_url      TEXT,
            ingested_at     TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS lng_export_terminals (
            id             BIGSERIAL PRIMARY KEY,
            project_name   TEXT NOT NULL,
            category       TEXT NOT NULL,   -- existing_or_under_construction | approved
            operator       TEXT,
            state          TEXT,
            trains         INTEGER,
            baseload_bcfd  NUMERIC(8,3),
            baseload_mtpa  NUMERIC(8,2),
            peak_bcfd      NUMERIC(8,3),
            peak_mtpa      NUMERIC(8,2),
            status         TEXT,
            in_service     TEXT,
            as_of          DATE,
            source_url     TEXT,
            ingested_at    TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
    """)
    # FULL unique index (not partial) so ON CONFLICT (project_name, category)
    # matches — the partial-index-vs-ON-CONFLICT trap has bitten this repo before.
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS ux_lng_export_terminals_name_cat
        ON lng_export_terminals (project_name, category)
    """)


# ══════════════════════════════════════════════════════════════════════
# 1) HENRY HUB daily spot — EIA v2 RNGWHHD
# ══════════════════════════════════════════════════════════════════════
def ingest_henry_hub(length=120, full=False):
    """Fetch RNGWHHD daily spot (sorted period DESC) and upsert into
    henry_hub_spot. Default pulls the latest 120 trading days (cheap daily
    top-up); full=True paginates the entire history (~7,400 rows).
    Returns (ok, parsed, debug) like the ISO queue ingestors."""
    debug = []
    api_key = os.environ.get("EIA_API_KEY")  # mirror eia_api.py (env at runtime)
    if not api_key:
        return False, None, "eia_api_key_missing (set EIA_API_KEY; no synthetic rows written)"
    if not NEON_URL:
        return False, None, "no_neon_url"

    PAGE = 5000
    per_page = PAGE if full else min(int(length), PAGE)
    rows = []
    offset = 0
    try:
        for _page in range(3 if full else 1):  # full history is ~7.4k rows; cap 15k
            params = {
                "frequency": "daily",
                "data[0]": "value",
                "facets[series][]": HH_SERIES,
                "sort[0][column]": "period",
                "sort[0][direction]": "desc",
                "length": per_page,
                "offset": offset,
            }
            url = f"{EIA_BASE}/{HH_ROUTE}?" + urllib.parse.urlencode(params)
            # key in a HEADER, never the query string
            body, st = _fetch(url, timeout=45,
                              headers={"X-Api-Key": api_key})
            if st != 200:
                return False, None, "; ".join(debug + [f"eia_http_{st}"])
            resp = (json.loads(body).get("response") or {})
            page = resp.get("data") or []
            debug.append(f"page{_page} http_{st} rows={len(page)} total={resp.get('total')}")
            rows.extend(page)
            if len(page) < per_page:
                break
            offset += per_page
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"eia_http_error_{e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"fetch_error: {type(e).__name__}: {str(e)[:100]}"])

    if not rows:
        return True, {}, "; ".join(debug + ["eia_empty_response"])

    upserts = []
    for r in rows:
        period = str(r.get("period") or "").strip()
        try:
            price = float(r.get("value"))
        except (TypeError, ValueError):
            continue  # EIA occasionally returns null/withheld — skip, never fabricate
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", period) or price <= 0:
            continue
        upserts.append((period, round(price, 3)))
    if not upserts:
        return True, {}, "; ".join(debug + ["no_parseable_rows"])

    try:
        with psycopg.connect(NEON_URL, autocommit=True) as conn, conn.cursor() as cur:
            _ensure_schema(cur)
            cur.executemany("""
                INSERT INTO henry_hub_spot (period, price_usd_mmbtu, source_url, ingested_at)
                VALUES (%s, %s, %s, NOW() ON CONFLICT DO NOTHING)
                ON CONFLICT (period) DO UPDATE
                SET price_usd_mmbtu = EXCLUDED.price_usd_mmbtu,
                    source_url      = EXCLUDED.source_url,
                    ingested_at     = NOW()
            """, [(p, v, HH_PUBLIC_URL) for p, v in upserts])
            cur.execute("SELECT COUNT(*), MAX(period) FROM henry_hub_spot")
            total, max_p = cur.fetchone()
    except Exception as e:
        return False, None, "; ".join(debug + [f"db_error: {type(e).__name__}: {str(e)[:120]}"])

    latest_period, latest_price = upserts[0]
    parsed = {
        "latest_period": latest_period,
        "latest_price_usd_mmbtu": latest_price,
        "rows_upserted": len(upserts),
        "table_rows": int(total),
        "table_max_period": max_p.isoformat() if max_p else None,
        "source_name": HH_SOURCE_LABEL,
        "source_url": HH_PUBLIC_URL,
    }
    return True, parsed, "; ".join(debug + [f"upserted={len(upserts)} latest={latest_period}@{latest_price}"])


# ══════════════════════════════════════════════════════════════════════
# 2) LNG EXPORT TERMINALS — EIA U.S. Liquefaction Capacity XLSX
# ══════════════════════════════════════════════════════════════════════
def _parse_lng_workbook(wb, debug):
    """Returns (as_of_date_or_None, [terminal dicts])."""
    # Release date from the Contents sheet ("Release Date:" label row).
    as_of = None
    if "Contents" in wb.sheetnames:
        for row in wb["Contents"].iter_rows(min_row=1, max_row=25, values_only=True):
            cells = [c for c in row if c is not None]
            for i, c in enumerate(cells):
                if isinstance(c, str) and "release date" in c.lower() and i + 1 < len(cells):
                    v = cells[i + 1]
                    if isinstance(v, datetime.datetime):
                        as_of = v.date()
                    break
            if as_of:
                break
    debug.append(f"release_date={as_of}")

    terminals = []

    # ── Sheet 1: Existing & Under Construction (per-TRAIN rows -> group by project)
    sheet1 = None
    for nm in wb.sheetnames:
        if "existing" in nm.lower():
            sheet1 = wb[nm]
            break
    if sheet1 is not None:
        # Header at row 2, units row 3, data from row 4. Columns (verified
        # 2026-07-01): 0 name, 1 train, 2 baseload Bcf/d, 3 baseload Mtpa,
        # 4 peak Bcf/d, 5 peak Mtpa, 6 status, 7 in-service, 9 state, 20 operator.
        projects = {}
        for row in sheet1.iter_rows(min_row=4, values_only=True):
            name = _cell_str(row[0] if len(row) > 0 else None)
            train = _cell_str(row[1] if len(row) > 1 else None)
            state = _cell_str(row[9] if len(row) > 9 else None)
            if not name or not train or not state:
                continue  # blank separators + Notes: footer rows
            if name.lower().startswith(("notes", "^", "source")):
                continue
            p = projects.setdefault(name, {
                "project_name": name, "category": "existing_or_under_construction",
                "operator": None, "state": state, "trains": 0,
                "baseload_bcfd": 0.0, "baseload_mtpa": 0.0,
                "peak_bcfd": 0.0, "peak_mtpa": 0.0,
                "statuses": [], "in_service": None,
            })
            tc = _train_count(train)
            if tc:
                p["trains"] += tc
            for key, col in (("baseload_bcfd", 2), ("baseload_mtpa", 3),
                             ("peak_bcfd", 4), ("peak_mtpa", 5)):
                v = _num(row[col] if len(row) > col else None)
                if v:
                    p[key] += v
            st = _clean_status(row[6] if len(row) > 6 else None)
            if st and st not in p["statuses"]:
                p["statuses"].append(st)
            if p["in_service"] is None:
                p["in_service"] = _cell_str(row[7] if len(row) > 7 else None)
            op = _cell_str(row[20] if len(row) > 20 else None)
            if op and not p["operator"]:
                p["operator"] = op
        for p in projects.values():
            p["status"] = " / ".join(p["statuses"]) or None
            del p["statuses"]
            terminals.append(p)
        debug.append(f"existing_projects={len(projects)}")

    # ── Sheet 2: Approved (per-project rows)
    sheet2 = wb["Approved"] if "Approved" in wb.sheetnames else None
    if sheet2 is not None:
        # Header at row 3, units row 4, data from row 5. Columns (verified
        # 2026-07-01): 0 name, 1 operator, 2 per-train Bcf/d, 3 per-train Mtpa,
        # 4 n trains, 5 total Bcf/d, 6 total Mtpa, 7 status, 8 state.
        n2 = 0
        for row in sheet2.iter_rows(min_row=5, values_only=True):
            name = _cell_str(row[0] if len(row) > 0 else None)
            operator = _cell_str(row[1] if len(row) > 1 else None)
            if not name or not operator:
                continue  # footnote / notes rows have no operator
            trains = _num(row[4] if len(row) > 4 else None)
            terminals.append({
                "project_name": name, "category": "approved",
                "operator": operator,
                "state": _cell_str(row[8] if len(row) > 8 else None),
                "trains": int(trains) if trains else None,
                "baseload_bcfd": _num(row[5] if len(row) > 5 else None),
                "baseload_mtpa": _num(row[6] if len(row) > 6 else None),
                "peak_bcfd": None, "peak_mtpa": None,
                "status": _clean_status(row[7] if len(row) > 7 else None) or None,
                "in_service": None,
            })
            n2 += 1
        debug.append(f"approved_projects={n2}")

    # Round the summed floats
    for t in terminals:
        for k in ("baseload_bcfd", "peak_bcfd"):
            if t.get(k):
                t[k] = round(t[k], 3)
        for k in ("baseload_mtpa", "peak_mtpa"):
            if t.get(k):
                t[k] = round(t[k], 2)
        for k in ("baseload_bcfd", "baseload_mtpa", "peak_bcfd", "peak_mtpa"):
            if t.get(k) == 0.0:
                t[k] = None
    return as_of, terminals


def ingest_lng_terminals():
    """Fetch + parse the EIA liquefaction-capacity workbook, then replace the
    lng_export_terminals table contents in one transaction (25-row reference
    table — full replace keeps it exactly in sync with EIA's current file)."""
    debug = []
    if not NEON_URL:
        return False, None, "no_neon_url"
    try:
        xlsx_bytes, st = _fetch(LNG_XLSX_URL, timeout=90, return_bytes=True)
        debug.append(f"xlsx http_{st} kb={len(xlsx_bytes)//1024}")
        if st != 200 or len(xlsx_bytes) < 10000:
            return False, None, "; ".join(debug + ["download_failed_or_small"])
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error_{e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"fetch_error: {type(e).__name__}: {str(e)[:100]}"])

    wb = _try_openpyxl(xlsx_bytes)
    if not wb:
        return False, None, "; ".join(debug + ["openpyxl_unavailable_or_bad_file"])
    try:
        as_of, terminals = _parse_lng_workbook(wb, debug)
    except Exception as e:
        return False, None, "; ".join(debug + [f"parse_error: {type(e).__name__}: {str(e)[:120]}"])
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # Sanity: EIA lists ~25 US projects. A tiny parse = format drift -> fail
    # loudly, keep the previous good rows (do NOT wipe the table).
    if len(terminals) < 10:
        return False, None, "; ".join(debug + [f"too_few_terminals={len(terminals)} (format drift?)"])

    try:
        with psycopg.connect(NEON_URL) as conn, conn.cursor() as cur:
            _ensure_schema(cur)
            cur.execute("DELETE FROM lng_export_terminals")
            for t in terminals:
                cur.execute("""
                    INSERT INTO lng_export_terminals
                        (project_name, category, operator, state, trains,
                         baseload_bcfd, baseload_mtpa, peak_bcfd, peak_mtpa,
                         status, in_service, as_of, source_url, ingested_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW() ON CONFLICT DO NOTHING)
                    ON CONFLICT (project_name, category) DO UPDATE
                    SET operator = EXCLUDED.operator, state = EXCLUDED.state,
                        trains = EXCLUDED.trains,
                        baseload_bcfd = EXCLUDED.baseload_bcfd,
                        baseload_mtpa = EXCLUDED.baseload_mtpa,
                        peak_bcfd = EXCLUDED.peak_bcfd, peak_mtpa = EXCLUDED.peak_mtpa,
                        status = EXCLUDED.status, in_service = EXCLUDED.in_service,
                        as_of = EXCLUDED.as_of, source_url = EXCLUDED.source_url,
                        ingested_at = NOW()
                """, (t["project_name"], t["category"], t["operator"], t["state"],
                      t["trains"] or None, t["baseload_bcfd"], t["baseload_mtpa"],
                      t["peak_bcfd"], t["peak_mtpa"], t["status"], t["in_service"],
                      as_of, LNG_XLSX_URL))
            conn.commit()
    except Exception as e:
        return False, None, "; ".join(debug + [f"db_error: {type(e).__name__}: {str(e)[:120]}"])

    total_bcfd = round(sum(t["baseload_bcfd"] or 0 for t in terminals
                           if t["category"] == "existing_or_under_construction"), 2)
    parsed = {
        "terminals": len(terminals),
        "existing_or_under_construction": sum(1 for t in terminals
                                              if t["category"] == "existing_or_under_construction"),
        "approved": sum(1 for t in terminals if t["category"] == "approved"),
        "existing_uc_baseload_bcfd": total_bcfd,
        "as_of": as_of.isoformat() if as_of else None,
        "source_name": LNG_SOURCE_LABEL,
        "source_url": LNG_XLSX_URL,
    }
    return True, parsed, "; ".join(debug + [f"replaced_rows={len(terminals)}"])


INGESTORS = {
    "henry_hub":     ingest_henry_hub,
    "lng_terminals": ingest_lng_terminals,
}


# ══════════════════════════════════════════════════════════════════════
# ingest endpoints (cron targets — same shape as /api/v1/iso-queue/ingest)
# ══════════════════════════════════════════════════════════════════════
@gas_feeds_bp.route("/feeds/ingest", methods=["GET", "POST"])
def ingest_all():
    if not is_valid_internal_key(request.headers.get("X-Internal-Key") or request.headers.get("X-Admin-Key")):
        return jsonify({"error": "unauthorized"}), 401
    started = datetime.datetime.now(datetime.timezone.utc)
    full = request.args.get("full") == "1"
    results = []
    for feed, fn in INGESTORS.items():
        try:
            ok, parsed, debug = fn(full=True) if (feed == "henry_hub" and full) else fn()
        except Exception as e:
            ok, parsed, debug = False, None, f"uncaught: {type(e).__name__}: {str(e)[:120]}"
        results.append({
            "feed": feed, "ok": ok,
            "parsed": parsed if ok else None,
            "debug": debug,
            "parser": FEED_STATUS.get(feed, "unknown"),
        })
    healthy = sum(1 for r in results if r["ok"])
    elapsed_ms = int((datetime.datetime.now(datetime.timezone.utc) - started).total_seconds() * 1000)
    return jsonify({
        "started_at": started.isoformat(),
        "elapsed_ms": elapsed_ms,
        "feeds_total": len(INGESTORS),
        "feeds_ok": healthy,
        "results": results,
        "note": ("henry_hub: EIA v2 RNGWHHD daily spot (?full=1 backfills entire "
                 "history). lng_terminals: EIA liquefaction-capacity XLSX full "
                 "replace. Basis hubs (Waha/DomSouth/Chicago/SoCal/Algonquin) are "
                 "paid-only (NGI/Argus/Platts) and intentionally absent."),
    }), 200 if healthy == len(INGESTORS) else 207


@gas_feeds_bp.route("/feeds/ingest/<feed>", methods=["GET", "POST"])
def ingest_one(feed):
    if not is_valid_internal_key(request.headers.get("X-Internal-Key") or request.headers.get("X-Admin-Key")):
        return jsonify({"error": "unauthorized"}), 401
    feed = feed.lower().replace("-", "_")
    fn = INGESTORS.get(feed)
    if not fn:
        return jsonify({"error": "unknown_feed", "feed": feed,
                        "available": list(INGESTORS.keys())}), 404
    kwargs = {}
    if feed == "henry_hub" and request.args.get("full") == "1":
        kwargs["full"] = True
    try:
        ok, parsed, debug = fn(**kwargs)
    except Exception as e:
        ok, parsed, debug = False, None, f"uncaught: {type(e).__name__}: {str(e)[:120]}"
    return jsonify({"feed": feed, "ok": ok, "parsed": parsed, "debug": debug,
                    "parser": FEED_STATUS.get(feed, "unknown")}), 200 if ok else 502


@gas_feeds_bp.route("/feeds/status", methods=["GET"])
def feeds_status():
    if not NEON_URL:
        return jsonify({"error": "no_neon_url"}), 500
    out = {"as_of": datetime.datetime.now(datetime.timezone.utc).isoformat(),
           "feeds": {}}
    try:
        with psycopg.connect(NEON_URL, autocommit=True) as conn, conn.cursor() as cur:
            try:
                cur.execute("""SELECT COUNT(*), MIN(period), MAX(period), MAX(ingested_at)
                               FROM henry_hub_spot""")
                n, mn, mx, ing = cur.fetchone()
                out["feeds"]["henry_hub"] = {
                    "rows": int(n), "oldest": mn.isoformat() if mn else None,
                    "latest": mx.isoformat() if mx else None,
                    "last_ingested_at": ing.isoformat() if ing else None,
                    "parser": FEED_STATUS["henry_hub"],
                }
            except Exception:
                out["feeds"]["henry_hub"] = {"rows": 0, "note": "table_absent (run ingest)"}
            try:
                cur.execute("""SELECT COUNT(*), MAX(as_of), MAX(ingested_at)
                               FROM lng_export_terminals""")
                n, aso, ing = cur.fetchone()
                out["feeds"]["lng_terminals"] = {
                    "rows": int(n), "as_of": aso.isoformat() if aso else None,
                    "last_ingested_at": ing.isoformat() if ing else None,
                    "parser": FEED_STATUS["lng_terminals"],
                }
            except Exception:
                out["feeds"]["lng_terminals"] = {"rows": 0, "note": "table_absent (run ingest)"}
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


# ══════════════════════════════════════════════════════════════════════
# public GET endpoints
# ══════════════════════════════════════════════════════════════════════
def _closest_price(cur, before_date):
    cur.execute("""SELECT period, price_usd_mmbtu FROM henry_hub_spot
                   WHERE period <= %s ORDER BY period DESC LIMIT 1""", (before_date,))
    return cur.fetchone()


@gas_feeds_bp.route("/henry-hub", methods=["GET"])
def henry_hub():
    """Latest Henry Hub daily spot + 1d/7d/30d trend + optional history.
    Params: days (history window, default 30, max 400)."""
    if not NEON_URL:
        return jsonify({"error": "no_neon_url"}), 500
    try:
        days = min(max(int(request.args.get("days", 30)), 1), 400)
    except ValueError:
        days = 30
    try:
        with psycopg.connect(NEON_URL, autocommit=True) as conn, conn.cursor() as cur:
            try:
                cur.execute("""SELECT period, price_usd_mmbtu FROM henry_hub_spot
                               ORDER BY period DESC LIMIT 1""")
                latest = cur.fetchone()
            except Exception:
                latest = None
            if not latest:
                return jsonify({
                    "data": None,
                    "note": ("No Henry Hub rows ingested yet — run POST "
                             "/api/v1/gas/feeds/ingest/henry_hub (requires EIA_API_KEY). "
                             "Values are never synthesized."),
                    "source": HH_SOURCE_LABEL,
                }), 200

            latest_period, latest_price = latest[0], float(latest[1])
            trend = {}
            # previous trading day
            cur.execute("""SELECT period, price_usd_mmbtu FROM henry_hub_spot
                           WHERE period < %s ORDER BY period DESC LIMIT 1""", (latest_period,))
            prev = cur.fetchone()
            for label, ref in (
                ("1d", prev),
                ("7d", _closest_price(cur, latest_period - datetime.timedelta(days=7))),
                ("30d", _closest_price(cur, latest_period - datetime.timedelta(days=30))),
            ):
                if ref and float(ref[1]) > 0:
                    ref_p, ref_v = ref[0], float(ref[1])
                    trend[label] = {
                        "ref_period": ref_p.isoformat(),
                        "ref_price_usd_mmbtu": round(ref_v, 3),
                        "change_usd": round(latest_price - ref_v, 3),
                        "change_pct": round(100.0 * (latest_price - ref_v) / ref_v, 1),
                    }
            cur.execute("""SELECT period, price_usd_mmbtu FROM henry_hub_spot
                           WHERE period >= %s ORDER BY period DESC""",
                        (latest_period - datetime.timedelta(days=days),))
            history = [{"period": r[0].isoformat(), "price_usd_mmbtu": float(r[1])}
                       for r in cur.fetchall()]
        return jsonify({
            "hub": "Henry Hub",
            "series": HH_SERIES,
            "units": "$/MMBtu",
            "latest": {"period": latest_period.isoformat(),
                       "price_usd_mmbtu": latest_price},
            "trend": trend,
            "history_days": days,
            "history": history,
            "as_of": latest_period.isoformat(),
            "source": HH_SOURCE_LABEL,
            "source_url": HH_PUBLIC_URL,
            "note": ("Henry Hub is the only free daily US gas spot price. Regional "
                     "basis hubs (Waha, Dominion South, Chicago, SoCal, Algonquin) "
                     "are paid-only indices (NGI/Argus/Platts) and are not served here."),
        })
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@gas_feeds_bp.route("/lng-terminals", methods=["GET"])
def lng_terminals():
    """US LNG export (liquefaction) terminals from the EIA capacity workbook.
    Params: state (2-letter), status (substring), category
    (existing_or_under_construction | approved)."""
    if not NEON_URL:
        return jsonify({"error": "no_neon_url"}), 500
    state = (request.args.get("state") or "").strip().upper()
    status = (request.args.get("status") or "").strip()
    category = (request.args.get("category") or "").strip().lower()
    where, params = [], []
    if state:
        where.append("UPPER(state) = %s")
        params.append(state)
    if status:
        where.append("status ILIKE %s")
        params.append(f"%{status}%")
    if category:
        where.append("category = %s")
        params.append(category)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    try:
        with psycopg.connect(NEON_URL, autocommit=True) as conn, conn.cursor() as cur:
            try:
                cur.execute(f"""
                    SELECT project_name, category, operator, state, trains,
                           baseload_bcfd, baseload_mtpa, peak_bcfd, peak_mtpa,
                           status, in_service, as_of, source_url
                    FROM lng_export_terminals
                    {where_sql}
                    ORDER BY baseload_bcfd DESC NULLS LAST, project_name
                """, params)
                rows = cur.fetchall()
            except Exception:
                rows = None
            if rows is None:
                return jsonify({
                    "data": None,
                    "note": ("No LNG terminal rows ingested yet — run POST "
                             "/api/v1/gas/feeds/ingest/lng_terminals."),
                    "source": LNG_SOURCE_LABEL,
                }), 200
        terminals = [{
            "project_name": r[0], "category": r[1], "operator": r[2],
            "state": r[3], "trains": r[4],
            "baseload_bcfd": float(r[5]) if r[5] is not None else None,
            "baseload_mtpa": float(r[6]) if r[6] is not None else None,
            "peak_bcfd": float(r[7]) if r[7] is not None else None,
            "peak_mtpa": float(r[8]) if r[8] is not None else None,
            "status": r[9], "in_service": r[10],
        } for r in rows]
        as_of = max((r[11] for r in rows if r[11]), default=None)
        totals = {}
        for t in terminals:
            c = t["category"]
            agg = totals.setdefault(c, {"terminals": 0, "baseload_bcfd": 0.0,
                                        "baseload_mtpa": 0.0})
            agg["terminals"] += 1
            agg["baseload_bcfd"] = round(agg["baseload_bcfd"] + (t["baseload_bcfd"] or 0), 2)
            agg["baseload_mtpa"] = round(agg["baseload_mtpa"] + (t["baseload_mtpa"] or 0), 1)
        return jsonify({
            "count": len(terminals),
            "terminals": terminals,
            "totals_by_category": totals,
            "as_of": as_of.isoformat() if as_of else None,
            "source": LNG_SOURCE_LABEL,
            "source_url": LNG_XLSX_URL,
            "note": ("Per-project (terminal/phase) capacity aggregated from EIA's "
                     "per-train rows; 'approved' projects are FERC/MARAD-approved, "
                     "not yet under construction. as_of = EIA workbook release date."),
        })
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500
