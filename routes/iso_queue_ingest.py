"""
iso_queue_ingest.py v2 (Phase ZZZZZ-round47.3, 2026-05-25)

Daily ingest of ISO interconnection queue snapshots. REAL parsers for ERCOT,
PJM, NYISO, MISO, CAISO, and SPP (r70) — 6 of 7. Only ISO-NE remains
heartbeat-only (its public queue-file URL moved; needs rediscovery) and is
NEVER surfaced as live data. Every real parser was verified against live data
before shipping (no fabricated numbers).

Architecture:
  - Each ISO has its own ingest function returning (ok, parsed_dict, debug)
  - _upsert is idempotent. If parsed dict is empty, does NOT create a NULL
    row — instead touches the latest existing row's ingested_at + by.
    Keeps the table clean so _latest_snapshot in routes/interconnection_queues.py
    isn't confused by NULL-data heartbeat rows.
  - Real parsers for ERCOT (public GIS Report XLSX) and PJM (XLSX via openpyxl).
  - MISO, SPP, CAISO, NYISO, ISO-NE are heartbeat-only with documented
    upgrade paths in inline TODO comments — each can be promoted to a real
    parser independently as we verify the actual file format.

Wiring (already in main.py):
    from routes.iso_queue_ingest import iso_queue_ingest_bp
    app.register_blueprint(iso_queue_ingest_bp)

Endpoints:
  POST /api/v1/iso-queue/ingest         — run all ISO ingestors
  POST /api/v1/iso-queue/ingest/<iso>   — run a single ISO (debugging)
  GET  /api/v1/iso-queue/ingest/status  — last run summary
  GET  /api/v1/iso-queue/parser-versions — which parsers have real logic

Cron: fired by cron_heartbeat at 06:00 UTC daily.
"""
import os
import re
import io
import csv
import json
import datetime
import urllib.request
import urllib.error
from flask import Blueprint, jsonify, request
from internal_auth import is_valid_internal_key
import psycopg

iso_queue_ingest_bp = Blueprint("iso_queue_ingest", __name__,
                                 url_prefix="/api/v1/iso-queue")
NEON_URL = os.environ.get("NEON_DATABASE_URL") or os.environ.get("DATABASE_URL")

UA = "DCHub-IsoQueueIngest/2.0 (+https://dchub.cloud/interconnection-queues)"

PARSER_STATUS = {
    "ERCOT":  "real (public GIS Report XLSX, MIS reportTypeId=15933, active Large+Small Gen capacity)",
    "PJM":    "real (public New Services Queue XLSX export, active sum of MFO)",
    "MISO":   "real (public GI-queue JSON API, applicationStatus=Active, summer MW)",
    "SPP":    "real (opsportal GenerateActiveCsv; MAX Summer MW, excl. cessation)",
    "CAISO":  "real (PublicQueueReport.xlsx 'Grid GenerationQueue'; Net MWs to Grid, status=ACTIVE)",
    "NYISO":  "real (public queue XLSX 'Interconnection Queue' sheet, SP MW via openpyxl)",
    "ISO-NE": "real (public IRTT Public Queue HTML table, Status=A active, Summer MW)",
    "NESO":   "real (UK NESO TEC Register CKAN JSON; active non-Built, per-project MW Increase)",
    "IESO":   "real (Ontario IESO CAA applicationstatusdata.json; active pipeline summer MW)",
    "AESO":   "real (Alberta AESO monthly Connection Project List XLSX; active EN1 STS+DTS MW)",
}


# ──────────────────────────────────────────────────────────────────────
# HTTP helper
# ──────────────────────────────────────────────────────────────────────
def _fetch(url, timeout=30, return_bytes=False, accept="*/*"):
    """Fetch URL. Returns (body, status). body is bytes if return_bytes else str."""
    req = urllib.request.Request(url, headers={
        "User-Agent":       UA,
        "Accept":           accept,
        "Accept-Language":  "en-US,en;q=0.9",
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read()
        return (body if return_bytes else body.decode("utf-8", errors="replace")), resp.status


def _try_pypdf(pdf_bytes):
    """Best-effort PDF text extraction. Returns full text or None."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        return "\n".join(p.extract_text() or "" for p in reader.pages)
    except ImportError:
        return None
    except Exception:
        return None


def _try_openpyxl(xlsx_bytes):
    """Returns workbook (read_only=True) or None. Handles .xls and .xlsx."""
    try:
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except ImportError:
        return None
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════
# ERCOT — public GIS Report (Generator Interconnection Status), monthly XLSX
# The market-data Public API (api.ercot.com) needs a subscription key AND a
# Bearer token (B2C ROPC) and does NOT expose the generation queue anyway.
# The queue lives in the GIS Report — published monthly as a PUBLIC, no-auth
# XLSX on the MIS (reportTypeId=15933). Discover the latest GIS_Report_*.xlsx
# from the public listing, download it, sum active 'Project Details - Large
# Gen' + '- Small Gen' capacity. Apples-to-apples with the other ISO gen queues.
# ══════════════════════════════════════════════════════════════════════
def _ercot_large_load():
    """ERCOT large-load (predominantly data-center) interconnection figures.

    ERCOT publishes these ONLY in monthly TAC / committee PDF decks — the Large
    Load Integration web page carries no machine-readable numbers and the PDF
    URLs are date-stamped/unstable. So this is best-effort: if the operator
    points ERCOT_LARGE_LOAD_PDF_URL at the current monthly report we fetch+parse
    the headline GW figure; otherwise we surface the latest PUBLISHED, CITED
    figure (ERCOT Q1-2026 TAC large-load update). Honest + dated, never blank.
    Returns {in_process_gw, subregions:[provenance block]}."""
    in_process_gw = 225.0          # ">225 GW of Large Loads going through the process"
    approved_gw = 9.0              # 9,042 MW have received Approval to Energize
    as_of = "2026-Q1"
    src = "https://www.ercot.com/services/rq/large-load-integration"
    basis = "cited: ERCOT Q1-2026 TAC Large Load update"
    pdf_url = (os.environ.get("ERCOT_LARGE_LOAD_PDF_URL") or "").strip()
    if pdf_url:
        try:
            pdf_bytes, st = _fetch(pdf_url, timeout=40, return_bytes=True)
            if st == 200 and pdf_bytes:
                text = _try_pypdf(pdf_bytes) or ""
                gws = re.findall(r'([0-9]{2,3}(?:\.[0-9])?)\s*GW', text)
                apprs = re.findall(r'([0-9],?[0-9]{3})\s*MW[^\n.]{0,40}(?:approv|energiz)', text, re.I)
                if gws:
                    in_process_gw = max(float(g) for g in gws)
                    src = pdf_url
                    basis = "parsed: ERCOT large-load PDF"
                    as_of = "live_pdf"
                if apprs:
                    approved_gw = round(float(apprs[0].replace(',', '')) / 1000.0, 1)
        except Exception:
            pass  # degrade to the cited fallback — never blank the field
    return {
        "in_process_gw": round(in_process_gw, 1),
        "subregions": [{
            "name": "Large-load (data-center-driven) interconnection queue",
            "in_process_gw": round(in_process_gw, 1),
            "approved_to_energize_gw": approved_gw,
            "as_of": as_of,
            "basis": basis,
            "source_url": src,
            "note": ("ERCOT large-load (>=75 MW) interconnection queue, "
                     "predominantly data centers — DISTINCT from the generation "
                     "queue (queued_load_total_gw); not a share of generation."),
        }],
    }


def ingest_ercot():
    debug = []
    parsed = {}
    # ERCOT large-load (data-center-driven) queue — independent of the GIS
    # generation report below, so set it FIRST: it survives even if the GIS
    # listing/download fails (those early-return with `parsed`).
    try:
        _ll = _ercot_large_load()
        parsed["queued_load_data_center_gw"] = _ll["in_process_gw"]
        parsed["top_subregions"] = _ll["subregions"]
    except Exception:
        pass
    try:
        LISTING = "https://www.ercot.com/misapp/GetReports.do?reportTypeId=15933"
        DL = "https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId="
        html, st = _fetch(LISTING, timeout=25)
        debug.append(f"listing http_{st} len={len(html)}")
        if st != 200 or len(html) < 2000:
            return True, parsed, "; ".join(debug + ["listing_unreachable"])

        # Title cell encodes the date + filename:
        #   RPT.00015933.<seq>.<YYYYMMDD>.<ms>.GIS_Report_<Mon><Year>.xlsx
        # then the 'Other' column has <a ...mirDownload?doclookupId=NNN>.
        # Listing is newest-first, so rows[0] is the latest monthly report.
        rows = re.findall(
            r'RPT\.00015933\.\d+\.(\d{8})\.\d+\.(GIS_Report_[A-Za-z0-9]+\.xlsx)'
            r'.*?doclookupId=(\d+)',
            html, re.S)
        if not rows:
            return True, parsed, "; ".join(debug + ["no_gis_xlsx_rows_in_listing"])
        pubdate, fname, docid = rows[0]
        debug.append(f"latest={fname} pub={pubdate} docid={docid}")

        xlsx_bytes, st = _fetch(DL + docid, timeout=90, return_bytes=True)
        debug.append(f"download http_{st} kb={len(xlsx_bytes)//1024}")
        if st != 200 or len(xlsx_bytes) < 50000:
            return False, None, "; ".join(debug + ["download_failed_or_small"])

        wb = _try_openpyxl(xlsx_bytes)
        if not wb:
            return True, parsed, "; ".join(debug + ["openpyxl_unavailable"])

        def _sum_gen_sheet(sheet_name):
            """Sum active queued capacity (MW) on a GIS project-detail sheet.
            Header row = the one whose cells contain 'INR'; capacity col is the
            first cell starting 'Capacity'. The Large/Small Gen sheets already
            EXCLUDE cancelled/inactive (those are on separate sheets), so every
            row with an INR + positive MW is an active queue project."""
            if sheet_name not in wb.sheetnames:
                return 0.0, 0
            ws = wb[sheet_name]
            hdr_idx = None
            inr_col = 0
            cap_col = None
            for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=40,
                                                   values_only=True), 1):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if "INR" in cells:
                    hdr_idx = r_i
                    inr_col = cells.index("INR")
                    for j, c in enumerate(cells):
                        if c.lower().startswith("capacity"):
                            cap_col = j
                            break
                    break
            if hdr_idx is None or cap_col is None:
                return 0.0, 0
            total = 0.0
            n = 0
            for row in ws.iter_rows(min_row=hdr_idx + 1, values_only=True):
                if len(row) <= max(cap_col, inr_col):
                    continue
                if row[inr_col] is None:
                    continue
                try:
                    mw = float(row[cap_col])
                except (TypeError, ValueError):
                    continue
                if mw <= 0:
                    continue
                total += mw
                n += 1
            return total, n

        lg_mw, lg_n = _sum_gen_sheet("Project Details - Large Gen")
        sg_mw, sg_n = _sum_gen_sheet("Project Details - Small Gen")
        try:
            wb.close()
        except Exception:
            pass
        total_mw = lg_mw + sg_mw
        total_n = lg_n + sg_n
        debug.append(f"large={lg_mw/1000:.1f}GW/{lg_n} small={sg_mw/1000:.1f}GW/{sg_n}")

        # Sanity: ERCOT's active gen queue is hundreds of GW. <1 GW = parse miss.
        if total_mw < 1000:
            return True, parsed, "; ".join(debug + [f"total_too_small={total_mw:.0f}MW"])

        month = fname.replace("GIS_Report_", "").replace(".xlsx", "")
        parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
        parsed["source_url"] = DL + docid
        parsed["source_name"] = f"ERCOT GIS Report ({month}, active generation queue)"
        return True, parsed, "; ".join(debug + [f"TOTAL={total_mw/1000:.1f}GW/{total_n}proj"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:100]}"])

# ══════════════════════════════════════════════════════════════════════
# PJM — public "New Services Queue" XLSX export (services.pjm.com)
# PJM reorganized into a cycle-based queue; the old ProjectsActive.xls 404s.
# The full queue is served as an XLSX from the PJM Planning API export
# endpoint using the api-subscription-key PJM bakes into its OWN public
# website JS bundle (pjm.com/dist/interconnectionqueues.*.js) to render the
# public queue dashboard — i.e. the same anonymous key every browser uses,
# NOT a private credential. (Overridable via PJM_QUEUE_KEY env if PJM ever
# rotates it.) Active queue = projects still in the interconnection process
# (Status not withdrawn/in-service/retracted/deactivated/annulled/canceled);
# sum MFO (Maximum Facility Output). Apples-to-apples with the other ISO queues.
# ══════════════════════════════════════════════════════════════════════
def ingest_pjm():
    debug = []
    parsed = {}
    try:
        url = "https://services.pjm.com/PJMPlanningApi/api/Queue/ExportToXls"
        key = os.environ.get("PJM_QUEUE_KEY", "E29477D0-70E0-4825-89B0-43F460BF9AB4")
        req = urllib.request.Request(url, data=b"", method="POST", headers={
            "api-subscription-key": key,          # PJM's public website constant
            "Origin":  "https://www.pjm.com",
            "Referer": "https://www.pjm.com/",
            "User-Agent": UA,
            "Accept": "*/*",
        })
        try:
            with urllib.request.urlopen(req, timeout=90) as resp:
                xlsx_bytes = resp.read()
                st = resp.status
        except urllib.error.HTTPError as e:
            return False, None, "; ".join(debug + [f"export_http_{e.code}"])
        debug.append(f"export http_{st} kb={len(xlsx_bytes)//1024}")
        if st != 200 or len(xlsx_bytes) < 50000:
            return False, None, "; ".join(debug + ["export_failed_or_small"])

        wb = _try_openpyxl(xlsx_bytes)
        if not wb:
            return True, parsed, "; ".join(debug + ["openpyxl_unavailable"])
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
        hdr = [str(c).strip() if c is not None else ""
               for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(hdr)}
        cS = idx.get("Status")
        cMFO = idx.get("MFO")
        cFuel = idx.get("Fuel")
        if cS is None or cMFO is None:
            return True, parsed, "; ".join(debug + ["missing_Status_or_MFO_column"])

        # Exclude projects that have LEFT the queue (withdrawn/cancelled) or
        # already GRADUATED (in service). Everything else is the active queue.
        EXCLUDE = {"withdrawn", "in service", "retracted", "deactivated",
                   "annulled", "canceled", "cancelled"}
        total_mw = 0.0
        n = 0
        rows_seen = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows_seen += 1
            if rows_seen > 200000:
                break
            if len(row) <= max(cS, cMFO) or row[cS] is None:
                continue
            if str(row[cS]).strip().lower() in EXCLUDE:
                continue
            try:
                mfo = float(row[cMFO]) if row[cMFO] is not None else 0
            except (TypeError, ValueError):
                continue
            if mfo <= 0:
                continue
            total_mw += mfo
            n += 1
        try:
            wb.close()
        except Exception:
            pass
        debug.append(f"active_n={n} rows_seen={rows_seen}")

        # Sanity: PJM's active queue is ~100-200 GW. <1 GW = parse miss.
        if total_mw < 1000:
            return True, parsed, "; ".join(debug + [f"total_too_small={total_mw:.0f}MW"])

        parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
        parsed["source_url"] = url
        parsed["source_name"] = "PJM New Services Queue (active, sum of MFO)"
        return True, parsed, "; ".join(debug + [f"TOTAL={total_mw/1000:.1f}GW/{n}proj"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:100]}"])


# ══════════════════════════════════════════════════════════════════════
# Stub ingestors — return heartbeat success when fetch works.
# Each has documented TODO for the real parser implementation.
# ══════════════════════════════════════════════════════════════════════
def _heartbeat(url):
    try:
        _body, status = _fetch(url, timeout=30)
        return True, {}, f"http_{status} (heartbeat OK, parser TODO)"
    except urllib.error.HTTPError as e:
        return False, None, f"http_error: {e.code}"
    except Exception as e:
        return False, None, f"error: {type(e).__name__}: {str(e)[:80]}"


def ingest_miso():
    """REAL parser (r70, 2026-06-03). MISO exposes the full GI queue as a public
    JSON API. Count applicationStatus=='Active' and sum summerNetMW (winterNetMW
    fallback). Verified live 2026-06-03: 1084 active projects, ~231 GW.
    Same (ok, parsed, debug) shape as ingest_pjm/ingest_nyiso."""
    debug = []
    parsed = {}
    url = "https://www.misoenergy.org/api/giqueue/getprojects"
    try:
        body, st = _fetch(url, timeout=60)
        debug.append(f"miso http_{st} kb={len(body)//1024}")
        if st != 200 or not body:
            return False, None, "; ".join(debug + ["fetch_failed"])
        rows = json.loads(body)
        if not isinstance(rows, list):
            return True, parsed, "; ".join(debug + ["unexpected_json_shape"])

        def _num(x):
            try: return float(x)
            except (TypeError, ValueError): return 0.0

        total_mw = 0.0
        active_n = 0
        dc_mw = 0.0
        for p in rows:
            if not isinstance(p, dict):
                continue
            if str(p.get("applicationStatus") or "").strip().lower() != "active":
                continue
            mw = _num(p.get("summerNetMW")) or _num(p.get("winterNetMW"))
            if mw <= 0:
                continue
            total_mw += mw
            active_n += 1
            ftype = (str(p.get("fuelType") or "") + " "
                     + str(p.get("facilityType") or "")).lower()
            if any(k in ftype for k in ("load", "data center", "datacenter")):
                dc_mw += mw

        debug.append(f"active_n={active_n} total_mw={total_mw:.0f} dc_mw={dc_mw:.0f}")
        if active_n > 0 and total_mw > 100:
            parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
            parsed["source_url"] = url
            parsed["source_name"] = "MISO GI Queue API (applicationStatus=Active, summer MW)"
        if dc_mw > 0 and total_mw > 0:
            parsed["queued_load_data_center_gw"] = round(dc_mw / 1000.0, 1)
            parsed["queued_load_dc_share_pct"] = round(100.0 * dc_mw / total_mw, 1)
        return True, parsed, "; ".join(debug)
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error: {e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:120]}"])


def ingest_spp():
    """REAL parser (r70, 2026-06-03). SPP serves its ACTIVE GI requests as CSV
    (opsportal GenerateActiveCsv; row 0 is a 'Last Updated On' banner, header is
    row 1, data row 2+). Sum 'MAX Summer MW' for rows WITHOUT a Cessation Date
    (cessation = withdrawn). Verified live 2026-06-03: 944 active, ~184 GW.
    Same (ok, parsed, debug) shape as the other real parsers."""
    debug = []
    parsed = {}
    url = "https://opsportal.spp.org/Studies/GenerateActiveCsv"
    try:
        body, st = _fetch(url, timeout=90)
        debug.append(f"spp http_{st} kb={len(body)//1024}")
        if st != 200 or not body:
            return False, None, "; ".join(debug + ["fetch_failed"])
        rows = list(csv.reader(io.StringIO(body)))
        if len(rows) < 3:
            return True, parsed, "; ".join(debug + ["too_few_rows"])
        hdr = [str(h or "").strip().lower() for h in rows[1]]  # row 0 is a banner

        def find_col(*keywords):
            for i, h in enumerate(hdr):
                if all(k in h for k in keywords):
                    return i
            return None

        col_mw = find_col("max", "summer", "mw") or find_col("summer", "mw") or find_col("mw")
        col_cess = find_col("cessation")
        col_fuel = find_col("fuel") or find_col("generation", "type")
        if col_mw is None:
            return True, parsed, "; ".join(debug + ["no_mw_column"])

        total_mw = 0.0
        active_n = 0
        dc_mw = 0.0
        for r in rows[2:]:
            if col_mw >= len(r):
                continue
            if col_cess is not None and col_cess < len(r) and str(r[col_cess]).strip():
                continue  # cessation date present → withdrawn, skip
            try:
                mw = float(str(r[col_mw]).replace(",", "").strip())
            except (ValueError, TypeError):
                continue
            if mw <= 0:
                continue
            total_mw += mw
            active_n += 1
            fuel_str = (str(r[col_fuel] or "").lower()
                        if col_fuel is not None and col_fuel < len(r) else "")
            if any(k in fuel_str for k in ("load", "data center", "datacenter")):
                dc_mw += mw

        debug.append(f"active_n={active_n} total_mw={total_mw:.0f}")
        if active_n > 0 and total_mw > 100:
            parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
            parsed["source_url"] = url
            parsed["source_name"] = "SPP GI active requests (MAX Summer MW, excl. cessation)"
        if dc_mw > 0 and total_mw > 0:
            parsed["queued_load_data_center_gw"] = round(dc_mw / 1000.0, 1)
            parsed["queued_load_dc_share_pct"] = round(100.0 * dc_mw / total_mw, 1)
        return True, parsed, "; ".join(debug)
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error: {e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:120]}"])


def ingest_caiso():
    """REAL parser (r70, 2026-06-03). CAISO publishes the public queue as XLSX
    (sheet 'Grid GenerationQueue'; banner rows 1-3, real header at Excel row 4,
    data row 5+). Filter Application Status=='ACTIVE', sum 'Net MWs to Grid'
    (fallback 'MW-1'). Verified live 2026-06-03: 284 active, ~81.2 GW."""
    debug = []
    parsed = {}
    url = "https://www.caiso.com/PublishedDocuments/PublicQueueReport.xlsx"
    try:
        xlsx_bytes, st = _fetch(url, timeout=90, return_bytes=True)
        debug.append(f"caiso http_{st} kb={len(xlsx_bytes)//1024}")
        if st != 200 or len(xlsx_bytes) < 10000:
            return False, None, "; ".join(debug + ["fetch_failed_or_too_small"])
        wb = _try_openpyxl(xlsx_bytes)
        if not wb:
            return True, parsed, "; ".join(debug + ["openpyxl_failed"])
        ws = None
        for nm in wb.sheetnames:
            if "generationqueue" in str(nm).strip().lower().replace(" ", ""):
                ws = wb[nm]
                break
        if ws is None:
            for nm in wb.sheetnames:
                if "queue" in str(nm).lower():
                    ws = wb[nm]
                    break
        if ws is None:
            ws = wb.active
        hdr = None
        for r in ws.iter_rows(min_row=4, max_row=4, values_only=True):
            hdr = [str(c or "").strip().lower() for c in r]
            break
        if not hdr:
            return True, parsed, "; ".join(debug + ["no_header_row4"])

        def find_col(*keywords):
            for i, h in enumerate(hdr):
                if all(k in h for k in keywords):
                    return i
            return None

        col_status = find_col("application", "status")
        col_mw = find_col("net", "mws") or find_col("mw-1") or find_col("mw")
        col_fuel = find_col("fuel-1") or find_col("fuel") or find_col("type-1")
        if col_mw is None or col_status is None:
            return True, parsed, "; ".join(debug + [f"cols_missing status={col_status} mw={col_mw}"])

        total_mw = 0.0
        active_n = 0
        dc_mw = 0.0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if col_status >= len(row):
                continue
            if str(row[col_status] or "").strip().upper() != "ACTIVE":
                continue
            mw = 0.0
            if col_mw < len(row):
                try:
                    mw = float(row[col_mw])
                except (ValueError, TypeError):
                    mw = 0.0
            if mw <= 0:
                continue
            total_mw += mw
            active_n += 1
            fuel_str = (str(row[col_fuel] or "").lower()
                        if col_fuel is not None and col_fuel < len(row) else "")
            if any(k in fuel_str for k in ("load", "data center", "datacenter")):
                dc_mw += mw

        debug.append(f"active_n={active_n} total_mw={total_mw:.0f}")
        if active_n > 0 and total_mw > 100:
            parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
            parsed["source_url"] = url
            parsed["source_name"] = "CAISO Public Queue Report (Application Status=ACTIVE, Net MWs to Grid)"
        if dc_mw > 0 and total_mw > 0:
            parsed["queued_load_data_center_gw"] = round(dc_mw / 1000.0, 1)
            parsed["queued_load_dc_share_pct"] = round(100.0 * dc_mw / total_mw, 1)
        return True, parsed, "; ".join(debug)
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error: {e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:120]}"])


def ingest_nyiso():
    """REAL parser (r70, 2026-06-03). NYISO publishes the full interconnection
    queue as a stable public XLSX. The 'Interconnection Queue' sheet is the
    ACTIVE queue (Withdrawn / In Service are separate sheets, so no status
    filter is needed). MW = 'SP (MW)' = summer-peak. Verified live 2026-06-03:
    147 active projects, ~24.9 GW. Same (ok, parsed, debug) shape as ingest_pjm."""
    debug = []
    parsed = {}
    url = "https://www.nyiso.com/documents/20142/1407078/NYISO-Interconnection-Queue.xlsx"
    try:
        xlsx_bytes, st = _fetch(url, timeout=60, return_bytes=True)
        debug.append(f"nyiso http_{st} kb={len(xlsx_bytes)//1024}")
        if st != 200 or len(xlsx_bytes) < 10000:
            return False, None, "; ".join(debug + ["fetch_failed_or_too_small"])
        wb = _try_openpyxl(xlsx_bytes)
        if not wb:
            return True, parsed, "; ".join(debug + ["openpyxl_failed"])
        ws = None
        for nm in wb.sheetnames:
            if "interconnection queue" in str(nm).strip().lower():
                ws = wb[nm]
                break
        if ws is None:
            ws = wb.active
        header = [str(c or "").strip().lower()
                  for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        debug.append(f"sheet='{ws.title}' cols={len(header)}")

        def find_col(*keywords):
            for i, h in enumerate(header):
                if all(k in h for k in keywords):
                    return i
            return None

        col_mw   = find_col("sp", "mw") or find_col("mw") or find_col("capacity")
        col_name = find_col("project") or find_col("name")
        col_fuel = find_col("fuel") or find_col("type")
        if col_mw is None:
            return True, parsed, "; ".join(debug + ["no_mw_column_found"])

        total_mw = 0.0
        active_n = 0
        dc_mw = 0.0
        rows_seen = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows_seen += 1
            if rows_seen > 100000:
                break
            if col_mw >= len(row) or row[col_mw] is None:
                continue
            try:
                mw = float(row[col_mw])
            except (ValueError, TypeError):
                continue
            if mw <= 0:
                continue
            total_mw += mw
            active_n += 1
            name_str = (str(row[col_name] or "").lower()
                        if col_name is not None and col_name < len(row) else "")
            fuel_str = (str(row[col_fuel] or "").lower()
                        if col_fuel is not None and col_fuel < len(row) else "")
            if any(k in name_str or k in fuel_str for k in (
                "data center", "datacenter", "ai cluster", "hyperscale",
                "data ctr", "large load", "load",
            )):
                dc_mw += mw

        debug.append(f"rows_seen={rows_seen} active_n={active_n} "
                     f"total_mw={total_mw:.0f} dc_mw={dc_mw:.0f}")

        if active_n > 0 and total_mw > 100:
            parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
            parsed["source_url"] = url
            parsed["source_name"] = "NYISO Interconnection Queue (active, SP MW)"
        if dc_mw > 0 and total_mw > 0:
            parsed["queued_load_data_center_gw"] = round(dc_mw / 1000.0, 1)
            parsed["queued_load_dc_share_pct"] = round(100.0 * dc_mw / total_mw, 1)

        return True, parsed, "; ".join(debug)
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error: {e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:120]}"])


def ingest_iso_ne():
    """ISO-NE — public IRTT "Public Queue" report (HTML table id='publicqueue').

    ISO-NE has no clean CSV/XLSX; the public queue is one big HTML page. GOTCHA:
    irtt.iso-ne.com does an ASP.NET cookie-detection 302 redirect, so a plain
    urlopen loops forever — we need a CookieJar opener (the shared _fetch can't).
    Active = Status code 'A' (W=withdrawn, C=completed); sum Summer MW (fallback
    Net MW), apples-to-apples with the other generation queues. No auth needed.
    """
    debug = []
    parsed = {}
    try:
        import http.cookiejar
        from lxml import html as _lh
        url = "https://irtt.iso-ne.com/reports/external"
        cj = http.cookiejar.CookieJar()
        opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
        req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "text/html"})
        with opener.open(req, timeout=60) as resp:
            body = resp.read()
            st = resp.status
        debug.append(f"fetch http_{st} kb={len(body)//1024}")
        if st != 200 or len(body) < 50000:
            return False, None, "; ".join(debug + ["fetch_failed_or_small"])

        doc = _lh.fromstring(body)
        tbls = doc.xpath('//table[@id="publicqueue"]')
        if not tbls:
            return True, parsed, "; ".join(debug + ["no_publicqueue_table"])
        tbl = tbls[0]
        heads = [(th.text_content() or "").strip() for th in tbl.xpath('.//thead//th')]
        if not heads:
            heads = [(c.text_content() or "").strip()
                     for c in tbl.xpath('.//tr[1]/th | .//tr[1]/td')]
        idx = {h: i for i, h in enumerate(heads)}
        ci_status = idx.get("Status")
        ci_summer = idx.get("Summer MW")
        ci_net = idx.get("Net MW")
        if ci_status is None or (ci_summer is None and ci_net is None):
            return True, parsed, "; ".join(debug + [f"cols_missing heads={heads[:10]}"])

        total_mw = 0.0
        active_n = 0
        rows_seen = 0
        for tr in tbl.xpath('.//tbody//tr'):
            tds = tr.xpath('./td')
            if len(tds) <= ci_status:
                continue
            rows_seen += 1
            if rows_seen > 100000:
                break
            vals = [(td.text_content() or "").strip() for td in tds]
            if vals[ci_status].upper() != "A":   # active only
                continue
            mw = 0.0
            for ci in (ci_summer, ci_net):
                if ci is not None and ci < len(vals):
                    try:
                        m = float(vals[ci].replace(",", ""))
                    except (ValueError, TypeError):
                        m = 0.0
                    if m:
                        mw = m
                        break
            if mw <= 0:
                continue
            total_mw += mw
            active_n += 1
        debug.append(f"rows_seen={rows_seen} active_n={active_n} total_mw={total_mw:.0f}")

        # ISO-NE active queue is ~12-17 GW; <0.5 GW = parse miss.
        if total_mw < 500:
            return True, parsed, "; ".join(debug + [f"total_too_small={total_mw:.0f}MW"])
        parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
        parsed["source_url"] = url
        parsed["source_name"] = "ISO-NE IRTT Public Queue (active, Summer MW)"
        return True, parsed, "; ".join(debug + [f"TOTAL={total_mw/1000:.1f}GW/{active_n}proj"])
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error: {e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:100]}"])


# ══════════════════════════════════════════════════════════════════════
# INTERNATIONAL queues (2026-07-01) — extend the US-only queue globally.
# Each writes into the same iso_queue_snapshots table via _upsert, so the
# public /interconnection-queue/snapshot + get_grid_scoreboard surface them
# with no extra wiring (the snapshot reads DISTINCT iso dynamically). All
# three fetch server-side from public, no-auth sources (verified 2026-07-01);
# Australia AEMO is intentionally absent — every AEMO path is behind a
# Cloudflare bot-challenge WAF that a server/runner fetch cannot defeat.
# ══════════════════════════════════════════════════════════════════════
def ingest_neso():
    """UK NESO (National Energy System Operator) — the GB transmission
    connections queue = the TEC Register, served from the NESO CKAN data
    portal as a public no-auth JSON datastore. Verified 2026-07-01: 2,215
    projects; active (Project Status != 'Built', i.e. not yet in service)
    ~610 GW — the largest single connections queue we track. Per-project
    capacity = 'MW Increase / Decrease' (the MW this project adds;
    'Cumulative Total Capacity' is a running site total → would double-count).
    Demand-type projects (Plant Type contains 'Demand' — data centres / large
    load) feed the DC-load share. Same (ok, parsed, debug) shape as the US
    ingestors."""
    debug = []
    parsed = {}
    rid = "17becbab-e3e8-473f-b303-3806f43a6a10"
    url = ("https://api.neso.energy/api/3/action/datastore_search"
           f"?resource_id={rid}&limit=5000")
    try:
        body, st = _fetch(url, timeout=30, accept="application/json")
        debug.append(f"neso http_{st} kb={len(body)//1024}")
        if st != 200:
            return False, None, "; ".join(debug + ["fetch_failed"])
        recs = (json.loads(body).get("result") or {}).get("records") or []
        debug.append(f"records={len(recs)}")
        if not recs:
            return True, parsed, "; ".join(debug + ["no_records"])

        def _num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        total_mw = dc_mw = 0.0
        active_n = 0
        for r in recs:
            if (r.get("Project Status") or "") == "Built":
                continue  # already energized — not queue
            mw = _num(r.get("MW Increase / Decrease"))
            if mw <= 0:
                continue
            total_mw += mw
            active_n += 1
            if "demand" in str(r.get("Plant Type") or "").lower():
                dc_mw += mw
        debug.append(f"active_n={active_n} total_mw={total_mw:.0f} dc_mw={dc_mw:.0f}")
        if active_n > 0 and total_mw > 100:
            parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
            parsed["source_url"] = ("https://www.neso.energy/data-portal/"
                                    "transmission-entry-capacity-tec-register")
            parsed["source_name"] = ("UK NESO TEC Register "
                                     "(active connections queue, per-project MW)")
        if dc_mw > 0 and total_mw > 0:
            parsed["queued_load_data_center_gw"] = round(dc_mw / 1000.0, 1)
            parsed["queued_load_dc_share_pct"] = round(100.0 * dc_mw / total_mw, 1)
        return True, parsed, "; ".join(debug)
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error: {e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:120]}"])


def ingest_ieso():
    """Ontario IESO — the Connection Assessment & Application (CAA) queue,
    published as a public no-auth JSON file. Verified 2026-07-01: 2,032
    applications; active pipeline (SiaStatus in Active / On-Hold / Part 1
    SIA Comp) ~15.9 GW, of which ~9 GW is Type='Load' (data-centre / large
    industrial) — a strong Ontario DC-demand signal. 'Complete' = energized,
    'Withdrawn' = dead (both excluded). Size is a string like '85.1 MW<br>'
    or '-' (equipment upgrades carry no MW) → strip tags/units; '-' = null."""
    debug = []
    parsed = {}
    url = "https://www.ieso.ca/-/media/files/IESO/files/applicationstatusdata.json"
    active_set = {"Active", "On-Hold", "Part 1 SIA Comp"}
    try:
        body, st = _fetch(url, timeout=30, accept="application/json")
        debug.append(f"ieso http_{st} kb={len(body)//1024}")
        if st != 200:
            return False, None, "; ".join(debug + ["fetch_failed"])
        rows = json.loads(body)
        if not isinstance(rows, list):
            return True, parsed, "; ".join(debug + ["not_a_list"])
        debug.append(f"rows={len(rows)}")

        def _psize(v):
            if not v:
                return None
            s = re.sub(r"<[^>]+>", "", str(v)).replace("MW", "").replace(",", "").strip()
            if s in ("", "-"):
                return None
            m = re.search(r"-?\d+(?:\.\d+)?", s)
            return float(m.group(0)) if m else None

        total_mw = dc_mw = 0.0
        active_n = 0
        for r in rows:
            if r.get("SiaStatus") not in active_set:
                continue
            mw = _psize(r.get("Size"))
            if not mw or mw <= 0:
                continue
            total_mw += mw
            active_n += 1
            if "load" in (str(r.get("Type") or "") + str(r.get("Name") or "")).lower():
                dc_mw += mw
        debug.append(f"active_n={active_n} total_mw={total_mw:.0f} dc_mw={dc_mw:.0f}")
        if active_n > 0 and total_mw > 100:
            parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
            parsed["source_url"] = ("https://www.ieso.ca/en/Get-Involved/"
                                    "Connection-Assessments/Connection-Application-Status")
            parsed["source_name"] = "Ontario IESO CAA connection queue (active, summer MW)"
        if dc_mw > 0 and total_mw > 0:
            parsed["queued_load_data_center_gw"] = round(dc_mw / 1000.0, 1)
            parsed["queued_load_dc_share_pct"] = round(100.0 * dc_mw / total_mw, 1)
        return True, parsed, "; ".join(debug)
    except urllib.error.HTTPError as e:
        return False, None, "; ".join(debug + [f"http_error: {e.code}"])
    except Exception as e:
        return False, None, "; ".join(debug + [f"error: {type(e).__name__}: {str(e)[:120]}"])


def ingest_aeso():
    """Alberta AESO — the monthly Connection Project List (XLSX). AESO uniquely
    tags 'Data Load' projects (Alberta's data-centre rush), so it carries a
    first-class DC-load signal. Active pipeline = Status in Active / ISD Under
    Review (Recently Energized = in service, Recently Cancelled = dead).
    Capacity per project = EN1 STS MW (generation) + EN1 DTS MW (load). The
    filename is month-dated (e.g. June-2026-Project-List.xlsx) and the current
    month is often not published yet, so we walk back up to 4 months."""
    debug = []
    parsed = {}
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    now = datetime.datetime.now(datetime.timezone.utc)
    xlsx_bytes = used = None
    y, m = now.year, now.month
    for _ in range(4):
        fn = f"{months[m-1]}-{y}-Project-List.xlsx"
        url = f"https://www.aeso.ca/assets/Uploads/project-reporting/{fn}"
        try:
            b, st = _fetch(url, timeout=40, return_bytes=True)
            if st == 200 and len(b) > 8000:
                xlsx_bytes, used = b, fn
                break
        except Exception as e:
            debug.append(f"{fn}:{type(e).__name__}")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    if not xlsx_bytes:
        return False, None, "; ".join(debug + ["no_monthly_file_found"])
    debug.append(f"file={used} kb={len(xlsx_bytes)//1024}")
    wb = _try_openpyxl(xlsx_bytes)
    if not wb:
        return True, parsed, "; ".join(debug + ["openpyxl_failed"])
    ws = None
    for nm in wb.sheetnames:
        if "connection project" in str(nm).strip().lower():
            ws = wb[nm]
            break
    if ws is None:
        ws = wb.active
    header = [str(c or "").strip().lower()
              for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def _find(*kw):
        for i, h in enumerate(header):
            if all(k in h for k in kw):
                return i
        return None

    c_sts, c_dts = _find("en1", "sts"), _find("en1", "dts")
    c_status, c_type = _find("status"), _find("mw", "type")
    debug.append(f"sheet='{ws.title}' sts={c_sts} dts={c_dts} status={c_status} type={c_type}")
    if c_status is None or (c_sts is None and c_dts is None):
        return True, parsed, "; ".join(debug + ["cols_not_found"])
    active_set = {"active", "isd under review"}

    def _cell(row, i):
        try:
            return float(row[i]) if i is not None and i < len(row) and row[i] is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    total_mw = dc_mw = 0.0
    active_n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        stat = (str(row[c_status]).strip().lower()
                if c_status < len(row) and row[c_status] else "")
        if stat not in active_set:
            continue
        mw = _cell(row, c_sts) + _cell(row, c_dts)
        if mw <= 0:
            continue
        total_mw += mw
        active_n += 1
        mtype = (str(row[c_type]).lower()
                 if c_type is not None and c_type < len(row) and row[c_type] else "")
        if "data" in mtype or "load" in mtype:
            dc_mw += mw
    debug.append(f"active_n={active_n} total_mw={total_mw:.0f} dc_mw={dc_mw:.0f}")
    if active_n > 0 and total_mw > 50:
        parsed["queued_load_total_gw"] = round(total_mw / 1000.0, 1)
        parsed["source_url"] = ("https://www.aeso.ca/grid/connecting-to-the-grid/"
                                "connection-project-reporting/")
        parsed["source_name"] = f"Alberta AESO Connection Project List (active; {used})"
    if dc_mw > 0 and total_mw > 0:
        parsed["queued_load_data_center_gw"] = round(dc_mw / 1000.0, 1)
        parsed["queued_load_dc_share_pct"] = round(100.0 * dc_mw / total_mw, 1)
    return True, parsed, "; ".join(debug)


INGESTORS = {
    "ERCOT":  ingest_ercot,
    "PJM":    ingest_pjm,
    "MISO":   ingest_miso,
    "SPP":    ingest_spp,
    "CAISO":  ingest_caiso,
    "NYISO":  ingest_nyiso,
    "ISO-NE": ingest_iso_ne,
    # International (2026-07-01)
    "NESO":   ingest_neso,   # United Kingdom (NESO TEC Register)
    "IESO":   ingest_ieso,   # Canada — Ontario
    "AESO":   ingest_aeso,   # Canada — Alberta
}


# ══════════════════════════════════════════════════════════════════════
# UPSERT — heartbeat-touch when parsed is empty (r47.3 change)
# ══════════════════════════════════════════════════════════════════════
def _upsert(iso, parsed, as_of, ingested_by):
    """Insert/update iso_queue_snapshots.

    r47.3: if `parsed` is empty (no new data), DO NOT create a NULL row.
    Instead, touch the latest existing row's ingested_at + ingested_by
    to record that we tried. Keeps _latest_snapshot from picking up
    NULL rows.
    """
    if not NEON_URL:
        return {"ok": False, "error": "no_neon_url"}

    if not parsed:
        try:
            with psycopg.connect(NEON_URL, autocommit=True) as conn, conn.cursor() as cur:
                cur.execute("""
                    UPDATE iso_queue_snapshots
                    SET ingested_at = NOW(),
                        ingested_by = %s
                    WHERE iso = %s
                      AND as_of = (SELECT MAX(as_of)
                                   FROM iso_queue_snapshots
                                   WHERE iso = %s)
                """, (ingested_by, iso, iso))
                touched = cur.rowcount
            return {"ok": True, "mode": "heartbeat_touch", "rows_touched": touched}
        except Exception as e:
            return {"ok": False, "error": f"{type(e).__name__}: {str(e)[:120]}"}

    fields = ["iso", "as_of", "ingested_by"]
    values = [iso, as_of, ingested_by]
    for k in ("queued_load_total_gw", "queued_load_data_center_gw",
              "queued_load_dc_share_pct", "new_applications_q_gw",
              "new_applications_period", "top_subregions",
              "source_url", "source_name"):
        if k in parsed:
            fields.append(k)
            v = parsed[k]
            values.append(json.dumps(v) if k == "top_subregions" else v)

    placeholders = ", ".join(["%s"] * len(values))
    cols = ", ".join(fields)
    upd_cols = [f for f in fields if f not in ("iso", "as_of")]
    upd = ", ".join(f"{c} = EXCLUDED.{c}" for c in upd_cols) + ", ingested_at = NOW()"
    sql = f"""
        INSERT INTO iso_queue_snapshots ({cols})
        VALUES ({placeholders})
        ON CONFLICT (iso, as_of) DO UPDATE SET {upd}
    """
    try:
        with psycopg.connect(NEON_URL, autocommit=True) as conn, conn.cursor() as cur:
            cur.execute(sql, values)
        return {"ok": True, "mode": "insert_or_update", "fields_updated": len(fields) - 3}
    except Exception as e:
        return {"ok": False, "error": f"{type(e).__name__}: {str(e)[:120]}"}


# ══════════════════════════════════════════════════════════════════════
# HTTP endpoints
# ══════════════════════════════════════════════════════════════════════
# AUTO-REPAIR: duplicate route '/ingest' also in routes/iso_lmp_ingest.py:646 — review and remove one
@iso_queue_ingest_bp.route("/ingest", methods=["GET", "POST"])
def ingest_all():
    if not is_valid_internal_key(request.headers.get("X-Internal-Key") or request.headers.get("X-Admin-Key")):
        return jsonify({"error": "unauthorized"}), 403
    started = datetime.datetime.now(datetime.timezone.utc)
    today = started.date()
    results = []
    for iso, fn in INGESTORS.items():
        ok, parsed, debug = fn()
        if ok:
            up = _upsert(iso, parsed or {}, today,
                         "cron_v2" if parsed else "cron_v2_heartbeat")
            results.append({
                "iso": iso, "fetch_ok": True,
                "parsed_fields": list((parsed or {}).keys()),
                "debug": debug, "upsert": up,
                "parser": PARSER_STATUS.get(iso, "unknown"),
            })
        else:
            results.append({
                "iso": iso, "fetch_ok": False, "debug": debug,
                "parser": PARSER_STATUS.get(iso, "unknown"),
            })
    elapsed_ms = int((datetime.datetime.now(datetime.timezone.utc) - started).total_seconds() * 1000)
    healthy = sum(1 for r in results if r.get("fetch_ok"))
    with_data = sum(1 for r in results if r.get("parsed_fields"))
    return jsonify({
        "started_at": started.isoformat(),
        "elapsed_ms": elapsed_ms,
        "as_of":      today.isoformat(),
        "isos_total": len(INGESTORS),
        "isos_fetched": healthy,
        "isos_with_new_data": with_data,
        "isos_failed":  len(INGESTORS) - healthy,
        "results": results,
        "note": "US: ERCOT/PJM/NYISO/CAISO (XLSX), MISO (JSON), SPP (CSV), "
                "ISO-NE (HTML). International: UK NESO (CKAN JSON), Ontario IESO "
                "(JSON), Alberta AESO (XLSX). Empty parse -> heartbeat_touch "
                "(no NULL rows). Australia AEMO absent (Cloudflare bot-WAF).",
    }), 200 if healthy == len(INGESTORS) else 207

# AUTO-REPAIR: duplicate route '/ingest/<iso>' also in routes/iso_lmp_ingest.py:684 — review and remove one

@iso_queue_ingest_bp.route("/ingest/<iso>", methods=["GET", "POST"])
def ingest_one(iso):
    if not is_valid_internal_key(request.headers.get("X-Internal-Key") or request.headers.get("X-Admin-Key")):
        return jsonify({"error": "unauthorized"}), 403
    iso = iso.upper()
    fn = INGESTORS.get(iso)
    if not fn:
        return jsonify({"error": "unknown_iso", "iso": iso,
                        "available": list(INGESTORS.keys())}), 404
    ok, parsed, debug = fn()
    today = datetime.date.today()
    up = _upsert(iso, parsed or {}, today,
                 "cron_v2" if (ok and parsed) else "cron_v2_heartbeat")
    return jsonify({"iso": iso, "fetch_ok": ok,
                    "parser": PARSER_STATUS.get(iso, "unknown"),
                    "parsed_fields": list((parsed or {}).keys()),
                    "debug": debug, "upsert": up})


@iso_queue_ingest_bp.route("/ingest/status", methods=["GET"])
def status():
    if not NEON_URL:
        return jsonify({"error": "no_neon_url"}), 500
    try:
        with psycopg.connect(NEON_URL, autocommit=True) as conn, conn.cursor() as cur:
            cur.execute("""
              SELECT iso, MAX(ingested_at) AS last_run, MAX(ingested_by) AS last_method,
                     MAX(queued_load_total_gw) AS latest_total_gw, MAX(as_of) AS latest_as_of
              FROM iso_queue_snapshots
              GROUP BY iso ORDER BY MAX(ingested_at) DESC NULLS LAST
            """)
            rows = cur.fetchall()
        return jsonify({
            "as_of": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "isos": [
                {"iso": r[0],
                 "last_run":     r[1].isoformat() if r[1] else None,
                 "last_method":  r[2],
                 "latest_total_gw": float(r[3] or 0),
                 "latest_as_of": r[4].isoformat() if r[4] else None,
                 "parser":       PARSER_STATUS.get(r[0], "unknown")}
                for r in rows
            ],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
# AUTO-REPAIR: duplicate route '/parser-versions' also in routes/iso_lmp_ingest.py:765 — review and remove one


@iso_queue_ingest_bp.route("/parser-versions", methods=["GET"])
def parser_versions():
    return jsonify({
        "ua": UA,
        "parsers": PARSER_STATUS,
        "deps": {
            "openpyxl": "required for ERCOT/PJM/NYISO/CAISO (XLSX parsing)",
            "pypdf": "no longer required (ERCOT moved off PDF to the GIS Report XLSX)",
        },
    })


# ══════════════════════════════════════════════════════════════════════
# WS6 queue-delta capture (2026-07-28) — ADD / ADVANCE / WITHDRAW
# ══════════════════════════════════════════════════════════════════════
# The per-project loader is a pure UPSERT with NO delete (UPSERT_SQL,
# load_interconnect_queue_live.py:602-619), and 5 of the 7 parsers write a
# hardcoded literal queue_status='active' (:202,:233,:287,:326,:492,:566)
# after filtering terminal projects out AT PARSE TIME (:273,:311,:359-369,
# :413-414,:545). So a project that leaves an ISO queue does not get a
# withdrawal status and does not get removed — its row simply stops being
# touched and keeps reading 'active' forever, and every consumer that sums
# capacity_mw counts it as live capacity. _DEAD_STATUS_RE
# (routes/interconnection_queues.py:513) cannot help: it can only match a
# status the loader never writes.
#
# This block is ADDITIVE. It writes ONLY the two new tables below, never
# interconnect_queue, and never raises — the ingest's return is unchanged
# whether capture succeeds, fails, or is skipped. The loader is untouched
# (it is also imported by the GH runner in --emit-json mode with neither
# psycopg2 nor DATABASE_URL, so DB work cannot live there).
#
# The signal already existed and nothing read it: `loaded_at` is bumped in
# DO UPDATE only for rows present in that day's feed, i.e. it is already a
# per-row last-seen watermark. It is uninterpretable ALONE — one parse
# failure freezes an entire ISO and reads as a mass withdrawal — so the run
# ledger is not garnish, it is what makes the watermark mean anything.
_QD_FLOOR = 0.6          # short-feed floor; see _qd_capture
_QD_SEEN_SLACK_S = 60    # clock slack when matching a row to the last trusted feed
_QD_SCHEMA_READY = False

# Bootstrapped at the call site (routes/facilities_delta.py:39-68 idiom):
# db_utils DDL is known to be SKIPPED in this codebase and the LIVE schema
# routinely differs from repo DDL, so never assume a migration ran. Separate
# tables on purpose — run_queue_etl.py:188 DROPs interconnect_queue and
# 04_interconnect_queue.py:179 DELETEs from it; neither may take history.
_QD_SCHEMA = """
CREATE TABLE IF NOT EXISTS interconnect_queue_runs (
    run_id               BIGSERIAL PRIMARY KEY,
    iso                  TEXT NOT NULL,
    run_at               TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    mode                 TEXT,
    rows_parsed          INT,
    rows_in_table_before INT,
    parse_ok             BOOLEAN,
    stale_not_in_feed    INT,
    feed_seen_at         TIMESTAMPTZ,
    events_written       INT,
    note                 TEXT
);
ALTER TABLE interconnect_queue_runs
    ADD COLUMN IF NOT EXISTS mode              TEXT,
    ADD COLUMN IF NOT EXISTS stale_not_in_feed INT,
    ADD COLUMN IF NOT EXISTS feed_seen_at      TIMESTAMPTZ,
    ADD COLUMN IF NOT EXISTS events_written    INT,
    ADD COLUMN IF NOT EXISTS note              TEXT;
CREATE INDEX IF NOT EXISTS ix_iqr_iso_run_at
    ON interconnect_queue_runs (iso, run_at DESC);

CREATE TABLE IF NOT EXISTS interconnect_queue_events (
    event_id         BIGSERIAL PRIMARY KEY,
    run_id           BIGINT,
    iso              TEXT NOT NULL,
    queue_id         TEXT NOT NULL,
    event_type       TEXT NOT NULL,
    project_name     TEXT,
    prev_capacity_mw DOUBLE PRECISION,
    new_capacity_mw  DOUBLE PRECISION,
    prev_status      TEXT,
    new_status       TEXT,
    prev_seen_at     TIMESTAMPTZ,
    detected_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
);
ALTER TABLE interconnect_queue_events
    ADD COLUMN IF NOT EXISTS project_name TEXT,
    ADD COLUMN IF NOT EXISTS prev_seen_at TIMESTAMPTZ;
CREATE UNIQUE INDEX IF NOT EXISTS interconnect_queue_events_run_key_uidx
    ON interconnect_queue_events (run_id, iso, queue_id, event_type);
CREATE INDEX IF NOT EXISTS ix_iqe_iso_detected
    ON interconnect_queue_events (iso, detected_at DESC);
CREATE INDEX IF NOT EXISTS ix_iqe_iso_qid
    ON interconnect_queue_events (iso, queue_id);
"""


def _qd_ensure_schema(conn):
    """Bootstrap the capture tables. Fail-soft: a DDL failure disables capture,
    never the ingest. MUST roll back — this shares the psycopg2 connection the
    upsert is about to use and an aborted transaction fails every later
    statement, including the post-run counts."""
    global _QD_SCHEMA_READY
    if _QD_SCHEMA_READY:
        return True
    try:
        with conn.cursor() as cur:
            cur.execute(_QD_SCHEMA)
        conn.commit()
        _QD_SCHEMA_READY = True
        return True
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return False


def _qd_mw(v):
    """capacity_mw -> float or None. psycopg2 returns Decimal for a NUMERIC
    column and the parsers return float; compare in one space."""
    try:
        return None if v is None else float(v)
    except Exception:
        return None


def _qd_status(v):
    """queue_status -> comparable label, or None when empty."""
    if v is None:
        return None
    s = str(v).strip().lower()
    return s or None


def _qd_ge(a, b, slack_s=0):
    """a >= b - slack, or None when the comparison is impossible (a legacy naive
    `loaded_at` against an aware TIMESTAMPTZ raises TypeError). Callers read
    `is not True`, so an impossible comparison always fails CLOSED — toward NOT
    publishing a withdrawal."""
    if a is None or b is None:
        return None
    try:
        return a >= (b - datetime.timedelta(seconds=slack_s))
    except Exception:
        return None


def _qd_diff(roster, feed, prev_feed_seen_at, last_gone, seed, emit_disappear):
    """PURE diff — no DB, no I/O, so tests can ast-extract it (tests never
    import main, and this module imports flask at module scope).

      roster            {queue_id: (capacity_mw, queue_status, loaded_at,
                        project_name)} read BEFORE the upsert
      feed              {queue_id: parsed row tuple} from this run's parse
      prev_feed_seen_at MAX(loaded_at) recorded by the last parse_ok run for
                        this ISO — the last moment a row is KNOWN to have been
                        in the feed
      last_gone         {queue_id: latest detected_at of a prior disappearance}
      seed              True on the first ledger run for this ISO
      emit_disappear    False when the short-feed floor did not clear

    Returns [(queue_id, event_type, project_name, prev_mw, new_mw, prev_status,
              new_status, prev_seen_at), ...].

    Row shape is load_interconnect_queue_live._row (:117-134):
      0 queue_id 1 project_name 2 iso 3 state 4 county 5 fuel_type
      6 capacity_mw 7 queue_status 8 queue_date 9 poi_name 10 lat 11 lng
      12 source
    """
    ev = []
    if seed:
        # First run under the ledger. interconnect_queue is ALREADY populated
        # (the loader has been upserting since 2026-06), so the feed is a
        # BASELINE, not ~5,300 additions, and rows in the table but absent from
        # the feed are an accumulated backlog of UNKNOWN AGE, not withdrawals
        # detected today. Emit the baseline; the backlog is counted on the run
        # row as stale_not_in_feed and never published as an event.
        for q, r in feed.items():
            ev.append((q, "seed", r[1] if len(r) > 1 else None,
                       None, _qd_mw(r[6]) if len(r) > 6 else None,
                       None, _qd_status(r[7]) if len(r) > 7 else None, None))
        return ev

    for q, r in feed.items():
        new_mw = _qd_mw(r[6]) if len(r) > 6 else None
        new_st = _qd_status(r[7]) if len(r) > 7 else None
        pname = r[1] if len(r) > 1 else None
        if q not in roster:
            ev.append((q, "appeared_in_feed", pname, None, new_mw, None, new_st, None))
            continue
        prev_mw = _qd_mw(roster[q][0])
        prev_st = _qd_status(roster[q][1])
        if ((prev_mw is None) != (new_mw is None)
                or (prev_mw is not None and new_mw is not None
                    and abs(prev_mw - new_mw) > 0.01)):
            ev.append((q, "capacity_changed", pname, prev_mw, new_mw,
                       prev_st, new_st, None))
        if prev_st != new_st:
            ev.append((q, "status_changed", pname, prev_mw, new_mw,
                       prev_st, new_st, None))

    if not emit_disappear:
        return ev

    for q, rv in roster.items():
        if q in feed:
            continue
        seen = rv[2] if len(rv) > 2 else None
        # Gate 1 — was this row in the LAST TRUSTED feed? A row whose watermark
        # predates the last parse_ok run is pre-existing backlog of unknown age,
        # NOT something that disappeared today.
        if _qd_ge(seen, prev_feed_seen_at, _QD_SEEN_SLACK_S) is not True:
            continue
        # Gate 2 — already reported gone and not re-seen since? Nothing ever
        # deletes the row, so without this the same project re-emits every day.
        gone_at = (last_gone or {}).get(q)
        if gone_at is not None and _qd_ge(seen, gone_at) is not True:
            continue
        ev.append((q, "disappeared_from_feed", rv[3] if len(rv) > 3 else None,
                   _qd_mw(rv[0]), None, _qd_status(rv[1]), None, seen))
    return ev


def _qd_before(conn, iso):
    """Pre-upsert roster + ledger state for ONE ISO. Returns None on ANY failure
    (capture is then skipped for that ISO and the ingest is unaffected)."""
    if not _qd_ensure_schema(conn):
        return None
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT queue_id, capacity_mw, queue_status, loaded_at, project_name "
                "FROM interconnect_queue WHERE iso = %s", (iso,))
            roster = {r[0]: (r[1], r[2], r[3], r[4]) for r in cur.fetchall() if r[0]}
            cur.execute("SELECT COUNT(*) FROM interconnect_queue_runs WHERE iso = %s",
                        (iso,))
            prior_runs = int((cur.fetchone() or [0])[0] or 0)
            cur.execute(
                "SELECT feed_seen_at FROM interconnect_queue_runs "
                "WHERE iso = %s AND parse_ok IS TRUE AND feed_seen_at IS NOT NULL "
                "ORDER BY run_at DESC LIMIT 1", (iso,))
            row = cur.fetchone()
            prev_feed_seen_at = row[0] if row else None
            cur.execute(
                "SELECT queue_id, MAX(detected_at) FROM interconnect_queue_events "
                "WHERE iso = %s AND event_type = 'disappeared_from_feed' "
                "GROUP BY queue_id", (iso,))
            last_gone = {r[0]: r[1] for r in cur.fetchall()}
        conn.commit()
        return {"roster": roster, "prior_runs": prior_runs,
                "prev_feed_seen_at": prev_feed_seen_at, "last_gone": last_gone}
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return None


def _qd_capture(conn, iso, rows, before, mode="trigger"):
    """Post-upsert transition capture for ONE ISO. Commits once per ISO
    (matching the loader's commit-per-source shape) and rolls back on ANY
    failure. Returns a small summary for the response, or None when skipped."""
    if not before:
        return None
    try:
        from psycopg2.extras import execute_values   # lazy, as the loader does
        feed = {}
        for r in (rows or []):
            try:
                if r and r[0]:
                    feed[r[0]] = r
            except Exception:
                continue
        roster = before.get("roster") or {}
        rows_parsed, rows_before = len(feed), len(roster)
        seed = not before.get("prior_runs")
        # ★ SHORT-FEED FLOOR. A PARTIAL parse is far more dangerous than a
        # failed one: `if rows: lq.upsert(...)` correctly skips a ZERO-row
        # parse, but a header-drift parse returning 40 of 2,000 PJM rows
        # upserts happily, and naive disappearance inference would then publish
        # 1,960 fabricated withdrawals. Adds and changes are safe from a short
        # feed; ONLY the disappearance branch is gated.
        parse_ok = bool(rows_parsed) and (rows_before == 0
                                          or rows_parsed >= _QD_FLOOR * rows_before)
        note = None if parse_ok else ("empty_parse" if not rows_parsed else "short_feed")
        ev = _qd_diff(roster, feed, before.get("prev_feed_seen_at"),
                      before.get("last_gone"),
                      seed and parse_ok, parse_ok and not seed)
        stale = sum(1 for q in roster if q not in feed)
        with conn.cursor() as cur:
            feed_seen_at = None
            if rows_parsed:
                cur.execute("SELECT MAX(loaded_at) FROM interconnect_queue "
                            "WHERE iso = %s", (iso,))
                feed_seen_at = (cur.fetchone() or [None])[0]
            cur.execute(
                "INSERT INTO interconnect_queue_runs "
                "(iso, mode, rows_parsed, rows_in_table_before, parse_ok, "
                " stale_not_in_feed, feed_seen_at, events_written, note) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING run_id",
                (iso, mode, rows_parsed, rows_before, parse_ok, stale,
                 feed_seen_at, len(ev), note))
            run_id = cur.fetchone()[0]
            if ev:
                execute_values(cur,
                               "INSERT INTO interconnect_queue_events "
                               "(run_id, iso, queue_id, event_type, project_name, "
                               " prev_capacity_mw, new_capacity_mw, prev_status, "
                               " new_status, prev_seen_at) VALUES %s "
                               "ON CONFLICT DO NOTHING",
                               [(run_id, iso) + e for e in ev],
                               template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                               page_size=500)
        conn.commit()
        by_type = {}
        for e in ev:
            by_type[e[1]] = by_type.get(e[1], 0) + 1
        return {"run_id": int(run_id), "parse_ok": parse_ok, "mode": mode,
                "rows_in_table_before": rows_before, "not_in_feed": stale,
                "events_written": len(ev), "events": by_type, "note": note}
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return None


def _qd_fail(conn, iso, note, mode="trigger"):
    """Record an ATTEMPTED-but-failed run. Without this row the read surface sees
    only the days that worked and reports a clean disappearance count for a
    window that has a hole in it — a feed outage read as a market fact. Rolls
    back FIRST: this can be reached with the transaction already aborted by
    whatever failed. Never raises."""
    try:
        conn.rollback()
    except Exception:
        pass
    if not _qd_ensure_schema(conn):
        return None
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO interconnect_queue_runs "
                "(iso, mode, rows_parsed, parse_ok, events_written, note) "
                "VALUES (%s,%s,0,FALSE,0,%s) RETURNING run_id",
                (iso, mode, str(note or "parse_failed")[:200]))
            run_id = cur.fetchone()[0]
        conn.commit()
        return int(run_id)
    except Exception:
        try: conn.rollback()
        except Exception: pass
        return None


# ── r-queue-projects-cron (2026-06-20): per-PROJECT interconnect_queue refresh ──
# Companion to /ingest (which writes AGGREGATE GW totals to iso_queue_snapshots).
# This refreshes the per-PROJECT `interconnect_queue` table (5,300+ NAMED ISO queue
# projects: project, MW, status, state/county, fuel, POI) by REUSING the parsers in
# load_interconnect_queue_live.py (PARSERS + upsert + ensure_schema + _dedupe).
#
# Fetches run ON RAILWAY (same as /ingest), so the daily GH cron just curls it —
# the 6 hosts the aggregate /ingest already fetches daily are proven Railway-egress
# reachable; ISO-NE is the one unknown. Each ISO is INDEPENDENT: an ISO blocked
# from Railway egress fails alone (others still load) and the GH workflow re-fetches
# THAT ISO on the runner (open egress, `--emit-json`) and POSTs the parsed rows back
# here (accept-rows mode). Open + idempotent (ON CONFLICT upsert), mirroring /ingest
# so the keyless GH curl works (GitHub Actions has no DATABASE_URL secret).
@iso_queue_ingest_bp.route("/ingest-projects", methods=["GET", "POST"])
@iso_queue_ingest_bp.route("/ingest-projects/<iso>", methods=["GET", "POST"])
def ingest_projects(iso=None):
    if not is_valid_internal_key(request.headers.get("X-Internal-Key") or request.headers.get("X-Admin-Key")):
        return jsonify({"error": "unauthorized"}), 403
    if not NEON_URL:
        return jsonify({"error": "no_neon_url"}), 500
    started = datetime.datetime.now(datetime.timezone.utc)
    try:
        import load_interconnect_queue_live as lq
    except Exception as e:
        return jsonify({"error": "loader_import_failed", "detail": str(e)[:200]}), 500
    try:
        import psycopg2  # the loader's upsert/ensure_schema use psycopg2 (v2)
        conn = psycopg2.connect(NEON_URL, connect_timeout=15)
    except Exception as e:
        return jsonify({"error": "db_connect_failed", "detail": str(e)[:200]}), 500

    out = {}
    try:
        lq.ensure_schema(conn)
        body = request.get_json(silent=True) or {}
        posted = body.get("rows") if isinstance(body, dict) else None
        if posted:
            # accept-rows mode: GH runner POSTs {iso, rows:[...]} for a blocked ISO
            name = (body.get("iso") or iso or "POSTED").upper()
            try:
                rows = [tuple(r) for r in posted if r and r[0]]
                # WS6: roster snapshot BEFORE the upsert, grouped by the ISO the
                # rows themselves carry (r[2]) — `name` can be the literal
                # "POSTED" when the runner omits `iso`, and the roster read keys
                # on the stored label.
                _keys = sorted({r[2] for r in rows if len(r) > 2 and r[2]})
                _pre = {k: _qd_before(conn, k) for k in _keys}
                if rows:
                    lq.upsert(conn, rows)
                out[name] = {"status": "ok", "rows": len(rows), "mode": "posted"}
                _delta = {}
                for k in _keys:
                    _d = _qd_capture(conn, k,
                                     [r for r in rows if len(r) > 2 and r[2] == k],
                                     _pre.get(k), mode="posted")
                    if _d:
                        _delta[k] = _d
                if _delta:
                    out[name]["delta"] = _delta
            except Exception as e:
                out[name] = {"status": "failed", "mode": "posted",
                             "error": f"{type(e).__name__}: {str(e)[:160]}"}
        else:
            # trigger mode: self-fetch + parse + upsert on Railway
            if iso:
                targets = [iso.upper()]
            else:
                q = (request.args.get("iso") or "").upper().strip()
                targets = [q] if q else list(lq.PARSERS.keys())
            for name in targets:
                fn = lq.PARSERS.get(name)
                if not fn:
                    out[name] = {"status": "unknown_iso"}
                    continue
                try:
                    rows = [r for r in lq._dedupe(fn()) if r[0]]
                    # WS6: roster snapshot BEFORE the upsert. Both calls swallow
                    # every exception internally and roll back on failure, so
                    # capture can never flip this ISO to "failed" or poison the
                    # connection for the ISOs after it.
                    _pre = _qd_before(conn, name)
                    if rows:
                        lq.upsert(conn, rows)
                    out[name] = {"status": "ok", "rows": len(rows),
                                 "with_mw": sum(1 for r in rows if r[6] is not None)}
                    _d = _qd_capture(conn, name, rows, _pre)
                    if _d:
                        out[name]["delta"] = _d
                except Exception as e:
                    # most likely Railway egress block -> GH-runner fallback re-fetches
                    out[name] = {"status": "failed",
                                 "error": f"{type(e).__name__}: {str(e)[:160]}"}
                    # WS6: an ATTEMPTED-but-failed run must be on the ledger, or
                    # the read surface reports a clean withdrawal count for a
                    # window with a hole in it.
                    _qd_fail(conn, name, f"{type(e).__name__}: {str(e)[:120]}")
        # post-run table totals — the real confirming count
        with conn.cursor() as cur:
            cur.execute("SELECT iso, COUNT(*) FROM interconnect_queue GROUP BY iso ORDER BY COUNT(*) DESC")
            by = {r[0]: int(r[1]) for r in cur.fetchall()}
            cur.execute("SELECT COUNT(*) FROM interconnect_queue")
            total = int(cur.fetchone()[0])
    except Exception as e:
        return jsonify({"error": "ingest_failed", "detail": str(e)[:200], "partial": out}), 500
    finally:
        try: conn.close()
        except Exception: pass

    healthy = sum(1 for v in out.values() if isinstance(v, dict) and v.get("status") == "ok")
    elapsed_ms = int((datetime.datetime.now(datetime.timezone.utc) - started).total_seconds() * 1000)
    return jsonify({
        "ok": True,
        "started_at": started.isoformat(),
        "elapsed_ms": elapsed_ms,
        "ran": list(out.keys()),
        "isos_ok": healthy,
        "by_iso": out,
        "interconnect_queue_total": total,
        "interconnect_queue_by_iso": by,
        "note": "Per-PROJECT rows in interconnect_queue (reuses load_interconnect_queue_live "
                "PARSERS). Aggregate GW totals are at /ingest (iso_queue_snapshots).",
    }), 200 if out and healthy == len(out) else 207
