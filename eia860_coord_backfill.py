"""
EIA-860 Coordinate Backfill for DC Hub Energy Discovery
========================================================
Downloads EIA-860 annual plant data (has lat/lng for every US plant),
matches against discovered_power_plants by plant ID, and updates coordinates.

Run on Replit: python3 eia860_coord_backfill.py
Requires: openpyxl, requests, psycopg2 (all should be available on Replit)
"""

import os
import sys
import requests
import tempfile
import time
from io import BytesIO

# Try openpyxl first, fall back to install hint
try:
    import openpyxl
except ImportError:
    print("❌ openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import psycopg2
except ImportError:
    print("❌ psycopg2 not installed. Run: pip install psycopg2-binary")
    sys.exit(1)

# =============================================================================
# CONFIG
# =============================================================================

DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    print("❌ DATABASE_URL not set in environment")
    sys.exit(1)

# EIA-860 annual data - try multiple URLs
EIA_860_URLS = [
    # Latest annual 860 (usually has a "2___Plant" or "Plant" sheet with lat/lng)
    "https://www.eia.gov/electricity/data/eia860/xls/eia8602024.zip",
    "https://www.eia.gov/electricity/data/eia860/xls/eia8602023.zip",
]

# Alternative: EIA-860M monthly (may have lat/lng in some versions)
EIA_860M_URLS = [
    "https://www.eia.gov/electricity/data/eia860m/xls/december_generator2026.xlsx",
    "https://www.eia.gov/electricity/data/eia860m/xls/october_generator2026.xlsx",
]


def download_file(url):
    """Download file to memory, return bytes"""
    print(f"   📥 Downloading: {url}")
    try:
        resp = requests.get(url, timeout=120, allow_redirects=True)
        if resp.status_code == 200:
            print(f"   ✅ Downloaded {len(resp.content):,} bytes")
            return resp.content
        elif resp.status_code == 301:
            redirect = resp.headers.get('Location', '')
            print(f"   ↪️ Redirect to: {redirect}")
            if redirect:
                return download_file(redirect)
        print(f"   ❌ HTTP {resp.status_code}")
        return None
    except Exception as e:
        print(f"   ❌ Download error: {e}")
        return None


def extract_xlsx_from_zip(zip_bytes):
    """Extract xlsx files from a zip archive"""
    import zipfile
    xlsx_files = {}
    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if name.endswith('.xlsx') or name.endswith('.xls'):
                print(f"   📁 Found: {name}")
                xlsx_files[name] = zf.read(name)
    return xlsx_files


def find_plant_sheet(wb):
    """Find the sheet with plant-level lat/lng data"""
    for sheet_name in wb.sheetnames:
        lower = sheet_name.lower()
        if 'plant' in lower and ('exist' in lower or '2__' in lower or '2___' in lower or 'operable' in lower):
            return sheet_name
        if sheet_name.lower() in ['plant', 'plants', '2___plant']:
            return sheet_name
    # Fallback: look for any sheet with lat/lng columns
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(cell.value or '').lower() for cell in ws[1]]
        if any('lat' in h for h in headers) and any('lon' in h or 'lng' in h for h in headers):
            return sheet_name
    return None


def parse_plant_coordinates(wb, sheet_name):
    """Parse plant ID → (lat, lng) from the plant sheet"""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return {}

    # Find header row (might not be row 1 — EIA often has title rows)
    header_row_idx = None
    for i, row in enumerate(rows):
        row_strs = [str(cell or '').lower().strip() for cell in row]
        if any('plant' in s and ('code' in s or 'id' in s) for s in row_strs):
            header_row_idx = i
            break
        if any('latitude' in s for s in row_strs):
            header_row_idx = i
            break

    if header_row_idx is None:
        # Try first row as header
        header_row_idx = 0

    headers = [str(cell or '').lower().strip() for cell in rows[header_row_idx]]
    print(f"   📋 Headers: {headers[:15]}...")

    # Find column indices
    plant_id_col = None
    lat_col = None
    lng_col = None
    name_col = None
    state_col = None

    for i, h in enumerate(headers):
        if ('plant' in h and ('code' in h or 'id' in h)) or h == 'plant code' or h == 'plantid' or h == 'plant id':
            plant_id_col = i
        elif 'latitude' in h or h == 'lat':
            lat_col = i
        elif 'longitude' in h or h == 'lng' or h == 'lon':
            lng_col = i
        elif 'plant name' in h or h == 'plant_name' or h == 'plantname':
            name_col = i
        elif h == 'state' or h == 'stateid':
            state_col = i

    if plant_id_col is None or lat_col is None or lng_col is None:
        print(f"   ❌ Couldn't find required columns. Plant ID col: {plant_id_col}, Lat: {lat_col}, Lng: {lng_col}")
        print(f"   Headers found: {headers}")
        return {}

    print(f"   ✅ Columns: plant_id={plant_id_col}, lat={lat_col}, lng={lng_col}, name={name_col}, state={state_col}")

    # Parse data rows
    coords = {}
    for row in rows[header_row_idx + 1:]:
        if len(row) <= max(plant_id_col, lat_col, lng_col):
            continue

        pid = row[plant_id_col]
        lat = row[lat_col]
        lng = row[lng_col]

        if pid is None or lat is None or lng is None:
            continue

        try:
            pid = str(int(float(pid)))
            lat = float(lat)
            lng = float(lng)
        except (ValueError, TypeError):
            continue

        # Sanity check: US coordinates
        if not (-180 <= lng <= -60 and 17 <= lat <= 72):
            continue

        name = str(row[name_col]) if name_col and len(row) > name_col else ''
        state = str(row[state_col]) if state_col and len(row) > state_col else ''

        coords[pid] = {'lat': lat, 'lng': lng, 'name': name, 'state': state}

    print(f"   📍 Parsed {len(coords):,} plants with coordinates")
    return coords


def backfill_coordinates(coords):
    """Update discovered_power_plants with lat/lng from EIA-860 lookup"""
    conn = psycopg2.connect(DATABASE_URL)
    c = conn.cursor()

    # Get all EIA-sourced plants missing coordinates
    c.execute("""
        SELECT id, name, state, market
        FROM discovered_power_plants
        WHERE (lat IS NULL OR lng IS NULL)
        AND id LIKE 'eia-%'
    """)
    missing = c.fetchall()
    print(f"\n🔍 Found {len(missing):,} EIA plants missing coordinates")

    updated = 0
    not_found = 0

    for row in missing:
        plant_id = row[0]  # e.g., "eia-6022"
        eia_pid = plant_id.replace('eia-', '')

        if eia_pid in coords:
            c.execute("""
                UPDATE discovered_power_plants
                SET lat = %s, lng = %s, source = COALESCE(source, '') || '+860', last_updated = NOW()::TEXT
                WHERE id = %s AND (lat IS NULL OR lng IS NULL)
            """, (coords[eia_pid]['lat'], coords[eia_pid]['lng'], plant_id))
            updated += 1
        else:
            not_found += 1

    # Also update any HIFLD plants that might match by name
    c.execute("""
        SELECT id, name, state, market
        FROM discovered_power_plants
        WHERE (lat IS NULL OR lng IS NULL)
        AND id NOT LIKE 'eia-%'
    """)
    non_eia_missing = c.fetchall()
    print(f"🔍 Found {len(non_eia_missing):,} non-EIA plants missing coordinates")

    # Build name lookup
    name_lookup = {}
    for pid, info in coords.items():
        if info['name']:
            name_lookup[info['name'].lower().strip()] = info

    name_matched = 0
    for row in non_eia_missing:
        plant_name = (row[1] or '').lower().strip()
        if plant_name in name_lookup:
            info = name_lookup[plant_name]
            c.execute("""
                UPDATE discovered_power_plants
                SET lat = %s, lng = %s, source = COALESCE(source, '') || '+860', last_updated = NOW()::TEXT
                WHERE id = %s AND (lat IS NULL OR lng IS NULL)
            """, (info['lat'], info['lng'], row[0]))
            name_matched += 1

    conn.commit()

    # Final stats
    c.execute("SELECT COUNT(*) FROM discovered_power_plants WHERE lat IS NOT NULL AND lng IS NOT NULL")
    with_coords = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM discovered_power_plants")
    total = c.fetchone()[0]

    conn.close()

    print(f"\n✅ BACKFILL COMPLETE")
    print(f"   Updated by plant ID: {updated:,}")
    print(f"   Updated by name match: {name_matched:,}")
    print(f"   Not found in EIA-860: {not_found:,}")
    print(f"   Total with coordinates: {with_coords:,} / {total:,} ({with_coords/total*100:.1f}%)")


def try_860m_xlsx(url):
    """Try parsing a monthly 860M xlsx for lat/lng"""
    data = download_file(url)
    if not data:
        return None

    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        print(f"   📊 Sheets: {wb.sheetnames}")

        # Check all sheets for lat/lng columns
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            first_rows = list(ws.iter_rows(max_row=3, values_only=True))
            for row in first_rows:
                row_strs = [str(cell or '').lower() for cell in row]
                if any('lat' in s for s in row_strs):
                    print(f"   🎯 Found lat/lng in sheet: {sheet_name}")
                    coords = parse_plant_coordinates(wb, sheet_name)
                    if coords:
                        return coords

        print(f"   ℹ️ No lat/lng columns found in this file")
        return None
    except Exception as e:
        print(f"   ❌ Parse error: {e}")
        return None


def try_860_zip(url):
    """Try parsing an annual 860 zip for lat/lng"""
    data = download_file(url)
    if not data:
        return None

    try:
        xlsx_files = extract_xlsx_from_zip(data)
        for name, content in xlsx_files.items():
            lower = name.lower()
            # The plant-level file is usually named like "2___Plant_Y2024.xlsx"
            if 'plant' in lower or '2___' in lower:
                print(f"   🎯 Opening: {name}")
                wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
                print(f"   📊 Sheets: {wb.sheetnames}")
                sheet = find_plant_sheet(wb)
                if sheet:
                    print(f"   🎯 Using sheet: {sheet}")
                    coords = parse_plant_coordinates(wb, sheet)
                    if coords:
                        return coords
        return None
    except Exception as e:
        print(f"   ❌ Zip parse error: {e}")
        return None


def main():
    print("=" * 60)
    print("🔋 DC Hub EIA-860 Coordinate Backfill")
    print("=" * 60)

    coords = None

    # Try annual EIA-860 first (most likely to have lat/lng)
    print("\n📦 Trying annual EIA-860 files...")
    for url in EIA_860_URLS:
        coords = try_860_zip(url)
        if coords:
            break

    # Fall back to monthly 860M
    if not coords:
        print("\n📦 Trying monthly EIA-860M files...")
        for url in EIA_860M_URLS:
            coords = try_860m_xlsx(url)
            if coords:
                break

    if not coords:
        print("\n❌ Could not find coordinates in any EIA file.")
        print("   Try downloading manually from: https://www.eia.gov/electricity/data/eia860/")
        print("   Look for the latest zip, extract the Plant xlsx, and check for Latitude/Longitude columns.")
        sys.exit(1)

    # Run the backfill
    backfill_coordinates(coords)


if __name__ == '__main__':
    main()
