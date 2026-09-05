#!/usr/bin/env python3
"""
load_interconnect_queue_live.py — PER-PROJECT ISO interconnection-queue loader.

Populates the `interconnect_queue` table with REAL per-project rows
(project, MW, status, location, fuel, POI) parsed live from the 7 US ISO
public queue feeds. This is the per-row companion to
routes/iso_queue_ingest.py, which only computes aggregate GW totals into
iso_queue_snapshots. The fetch URLs / parse patterns (PJM public key,
ISO-NE cookie jar, ERCOT GIS listing scrape, CAISO row-4 header, etc.) are
reused from that proven module.

Run it from inside your dchub-backend checkout (the directory linked to the
Railway project) so `railway run` injects the prod DATABASE_URL; egress then
comes from your workstation's network rather than Railway's:

    railway run python3 load_interconnect_queue_live.py            # all ISOs
    railway run python3 load_interconnect_queue_live.py NYISO CAISO # subset

Schema (must match exactly — verified live):
  interconnect_queue(id serial, queue_id text, project_name text, iso text,
    state text, county text, fuel_type text, capacity_mw numeric,
    queue_status text, queue_date date, poi_name text, lat numeric,
    lng numeric, source text, loaded_at timestamptz)

Idempotent: ensures a UNIQUE(iso, queue_id) index, then UPSERTs (ON CONFLICT).
A parser failure on one ISO never aborts the others. Geocode left NULL.
"""
import os
import io
import sys
import csv
import json
import re
import datetime
import urllib.request
import urllib.error

DATABASE_URL = os.environ.get("DATABASE_URL") or os.environ.get("NEON_DATABASE_URL")
# r-queue-projects-cron (2026-06-20): DATABASE_URL + psycopg2 are needed ONLY for
# the DB-writing path (main() DB run + upsert()), where psycopg2 is imported
# lazily. This keeps the module IMPORT-SAFE for two new callers:
#   (a) routes/iso_queue_ingest.py importing PARSERS/upsert/ensure_schema/_dedupe
#       for the daily /api/v1/iso-queue/ingest-projects route, and
#   (b) the `--emit-json` parse-only path run on a GitHub Actions runner that has
#       NEITHER DATABASE_URL nor psycopg2 — the GH-runner-fetch->POST fallback for
#       any ISO whose feed is blocked from Railway egress.

UA = "DCHub-IsoQueueIngest/2.0 (+https://dchub.cloud/interconnection-queues)"


# ──────────────────────────────────────────────────────────────────────
# helpers
# ──────────────────────────────────────────────────────────────────────
def _fetch(url, timeout=90, return_bytes=False, method="GET", data=None, headers=None):
    h = {"User-Agent": UA, "Accept": "*/*", "Accept-Language": "en-US,en;q=0.9"}
    if headers:
        h.update(headers)
    req = urllib.request.Request(url, data=data, method=method, headers=h)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read()
        return (body if return_bytes else body.decode("utf-8", "replace")), resp.status


def _wb(xlsx_bytes):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)


def _s(v, maxlen=None):
    """Clean a cell to a stripped string or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "None", "nan", "NaN", "N/A", "NA", "TBD", "-"):
        return None
    return s[:maxlen] if maxlen else s


def _mw(v):
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", "").strip())
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def _date(v):
    """Parse many date shapes (datetime, m/d/Y, YYYY-MM-DD, MM-YYYY) → date|None."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    if not s or s in ("None", "nan", "N/A", "NA", "TBD", "-"):
        return None
    # ISO datetime
    s2 = s.split("T")[0].split(" ")[0]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.datetime.strptime(s2, fmt).date()
        except ValueError:
            pass
    # MM-YYYY or MM/YYYY (proposed COD style) → first of month
    m = re.match(r"^(\d{1,2})[-/](\d{4})$", s)
    if m:
        try:
            return datetime.date(int(m.group(2)), int(m.group(1)), 1)
        except ValueError:
            return None
    return None


def _row(queue_id, project_name, iso, state, county, fuel_type, capacity_mw,
         queue_status, queue_date, poi_name, source):
    """Build a table-shaped tuple. lat/lng left NULL. queue_id required."""
    return (
        _s(queue_id, 120) or "",
        _s(project_name, 300),
        iso,
        _s(state, 8),
        _s(county, 160),
        _s(fuel_type, 160),
        capacity_mw,
        _s(queue_status, 120),
        queue_date,
        _s(poi_name, 300),
        None,            # lat
        None,            # lng
        source,
    )


def _hdr_index(header_cells):
    """Return {lowercased-collapsed-header: col_index}. Collapses newlines/spaces."""
    out = {}
    for i, c in enumerate(header_cells):
        key = re.sub(r"\s+", " ", str(c or "").strip().lower())
        if key and key not in out:
            out[key] = i
    return out


def _find(idx, *cands):
    """Find a column index by trying exact then substring matches."""
    for c in cands:
        if c in idx:
            return idx[c]
    for c in cands:
        for k, i in idx.items():
            if c in k:
                return i
    return None


# ══════════════════════════════════════════════════════════════════════
# NYISO — Interconnection Queue (gen) + Load Projects sheets
# ══════════════════════════════════════════════════════════════════════
def parse_nyiso():
    url = "https://www.nyiso.com/documents/20142/1407078/NYISO-Interconnection-Queue.xlsx"
    src = "NYISO Interconnection Queue (live XLSX)"
    b, st = _fetch(url, timeout=90, return_bytes=True)
    if st != 200 or len(b) < 10000:
        raise RuntimeError(f"fetch http_{st} small={len(b)}")
    wb = _wb(b)
    rows = []
    TYPEFUEL = {"S": "Solar", "W": "Wind", "ES": "Energy Storage", "L": "Load",
                "G": "Gas", "N": "Nuclear", "H": "Hydro", "CC": "Combined Cycle",
                "CT": "Combustion Turbine", "ST": "Steam"}
    # --- generation queue sheet (header row 1, active queue) ---
    if "Interconnection Queue" in wb.sheetnames:
        ws = wb["Interconnection Queue"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        idx = _hdr_index(header)
        c_q = _find(idx, "queue pos.", "queue pos", "queue number", "queue")
        c_name = _find(idx, "project name", "project: project name")
        c_dev = _find(idx, "developer/interconnect", "developer name")
        c_mw = _find(idx, "sp (mw)", "peak mw load", "mw")
        c_date = _find(idx, "date of ir", "ir submission date")
        c_fuel = _find(idx, "type/ fuel", "type/fuel")
        c_county = _find(idx, "county")
        c_state = _find(idx, "state")
        c_poi = _find(idx, "points of interconnect", "point of interconnect")
        c_rec = _find(idx, "record type")
        for r in ws.iter_rows(min_row=2, values_only=True):
            qid = _s(r[c_q]) if c_q is not None and c_q < len(r) else None
            if not qid:
                continue
            name = (_s(r[c_name]) if c_name is not None and c_name < len(r) else None) \
                or (_s(r[c_dev]) if c_dev is not None and c_dev < len(r) else None)
            ftc = _s(r[c_fuel]) if c_fuel is not None and c_fuel < len(r) else None
            fuel = TYPEFUEL.get(ftc, ftc) if ftc else None
            rows.append(_row(
                f"NYISO-{qid}", name, "NYISO",
                _s(r[c_state]) if c_state is not None and c_state < len(r) else "NY",
                _s(r[c_county]) if c_county is not None and c_county < len(r) else None,
                fuel,
                _mw(r[c_mw]) if c_mw is not None and c_mw < len(r) else None,
                "active",
                _date(r[c_date]) if c_date is not None and c_date < len(r) else None,
                _s(r[c_poi]) if c_poi is not None and c_poi < len(r) else None,
                src,
            ))
    # --- load projects sheet (data-center & large load) ---
    if "Load Projects" in wb.sheetnames:
        ws = wb["Load Projects"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        idx = _hdr_index(header)
        c_q = _find(idx, "queue number", "queue")
        c_name = _find(idx, "project: project name", "project name")
        c_dev = _find(idx, "developer name")
        c_mw = _find(idx, "peak mw load", "mw")
        c_date = _find(idx, "ir submission date")
        c_fuel = _find(idx, "type/fuel", "type/ fuel")
        c_county = _find(idx, "county")
        c_state = _find(idx, "state")
        c_poi = _find(idx, "points of interconnect")
        for r in ws.iter_rows(min_row=2, values_only=True):
            qid = _s(r[c_q]) if c_q is not None and c_q < len(r) else None
            if not qid:
                continue
            name = (_s(r[c_name]) if c_name is not None and c_name < len(r) else None) \
                or (_s(r[c_dev]) if c_dev is not None and c_dev < len(r) else None)
            rows.append(_row(
                f"NYISO-{qid}", name, "NYISO",
                _s(r[c_state]) if c_state is not None and c_state < len(r) else "NY",
                _s(r[c_county]) if c_county is not None and c_county < len(r) else None,
                "Load",
                _mw(r[c_mw]) if c_mw is not None and c_mw < len(r) else None,
                "active",
                _date(r[c_date]) if c_date is not None and c_date < len(r) else None,
                _s(r[c_poi]) if c_poi is not None and c_poi < len(r) else None,
                src,
            ))
    return rows


# ══════════════════════════════════════════════════════════════════════
# CAISO — Grid GenerationQueue sheet (header at Excel row 4, data 5+)
# ══════════════════════════════════════════════════════════════════════
def parse_caiso():
    url = "https://www.caiso.com/PublishedDocuments/PublicQueueReport.xlsx"
    src = "CAISO Public Queue Report (live XLSX)"
    b, st = _fetch(url, timeout=90, return_bytes=True)
    if st != 200 or len(b) < 10000:
        raise RuntimeError(f"fetch http_{st} small={len(b)}")
    wb = _wb(b)
    ws = None
    for nm in wb.sheetnames:
        if "generationqueue" in str(nm).lower().replace(" ", ""):
            ws = wb[nm]
            break
    if ws is None:
        raise RuntimeError("no Grid GenerationQueue sheet")
    header = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    idx = _hdr_index(header)
    c_name = _find(idx, "project name")
    c_q = _find(idx, "queue position")
    c_date = _find(idx, "queue date", "interconnection request receive date")
    c_status = _find(idx, "application status")
    c_fuel = _find(idx, "fuel-1", "fuel")
    c_type = _find(idx, "type-1", "type")
    c_mw = _find(idx, "net mws to grid", "mw-1")
    c_county = _find(idx, "county")
    c_state = _find(idx, "state")
    c_poi = _find(idx, "station or transmission line")
    rows = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if c_status is not None and c_status < len(r):
            if str(r[c_status] or "").strip().upper() != "ACTIVE":
                continue
        qid = _s(r[c_q]) if c_q is not None and c_q < len(r) else None
        name = _s(r[c_name]) if c_name is not None and c_name < len(r) else None
        if not qid and not name:
            continue
        fuel = (_s(r[c_fuel]) if c_fuel is not None and c_fuel < len(r) else None) \
            or (_s(r[c_type]) if c_type is not None and c_type < len(r) else None)
        rows.append(_row(
            f"CAISO-{qid}" if qid else f"CAISO-{name}", name, "CAISO",
            _s(r[c_state]) if c_state is not None and c_state < len(r) else "CA",
            _s(r[c_county]) if c_county is not None and c_county < len(r) else None,
            fuel,
            _mw(r[c_mw]) if c_mw is not None and c_mw < len(r) else None,
            "active",
            _date(r[c_date]) if c_date is not None and c_date < len(r) else None,
            _s(r[c_poi]) if c_poi is not None and c_poi < len(r) else None,
            src,
        ))
    return rows


# ══════════════════════════════════════════════════════════════════════
# MISO — public GI queue JSON API
# ══════════════════════════════════════════════════════════════════════
def parse_miso():
    url = "https://www.misoenergy.org/api/giqueue/getprojects"
    src = "MISO GI Queue API (live JSON)"
    body, st = _fetch(url, timeout=90)
    if st != 200 or not body:
        raise RuntimeError(f"fetch http_{st}")
    data = json.loads(body)
    if not isinstance(data, list):
        raise RuntimeError("unexpected json shape")
    rows = []
    for p in data:
        if not isinstance(p, dict):
            continue
        if str(p.get("applicationStatus") or "").strip().lower() != "active":
            continue
        qid = _s(p.get("projectNumber")) or _s(p.get("id"))
        if not qid:
            continue
        mw = _mw(p.get("summerNetMW")) or _mw(p.get("winterNetMW"))
        fuel = _s(p.get("fuelType")) or _s(p.get("facilityType"))
        rows.append(_row(
            f"MISO-{qid}",
            _s(p.get("transmissionOwner")) or _s(p.get("poiName")) or qid,
            "MISO",
            _s(p.get("state")),
            _s(p.get("county")),
            fuel,
            mw,
            "active",
            _date(p.get("inService")),
            _s(p.get("poiName")),
            src,
        ))
    return rows


# ══════════════════════════════════════════════════════════════════════
# PJM — New Services Queue XLSX (public api-subscription-key)
# ══════════════════════════════════════════════════════════════════════
def parse_pjm():
    url = "https://services.pjm.com/PJMPlanningApi/api/Queue/ExportToXls"
    src = "PJM New Services Queue (live XLSX)"
    key = os.environ.get("PJM_QUEUE_KEY", "E29477D0-70E0-4825-89B0-43F460BF9AB4")
    b, st = _fetch(url, timeout=120, return_bytes=True, method="POST", data=b"",
                   headers={"api-subscription-key": key,
                            "Origin": "https://www.pjm.com",
                            "Referer": "https://www.pjm.com/"})
    if st != 200 or len(b) < 50000:
        raise RuntimeError(f"export http_{st} small={len(b)}")
    wb = _wb(b)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = _hdr_index(header)
    c_id = _find(idx, "project id")
    c_name = _find(idx, "commercial name", "name")
    c_state = _find(idx, "state")
    c_county = _find(idx, "county")
    c_status = _find(idx, "status")
    c_mw = _find(idx, "mfo")
    c_fuel = _find(idx, "fuel")
    c_date = _find(idx, "submitted date")
    EXCLUDE = {"withdrawn", "in service", "retracted", "deactivated",
               "annulled", "canceled", "cancelled"}
    rows = []
    seen = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        seen += 1
        if seen > 200000:
            break
        status_raw = _s(r[c_status]) if c_status is not None and c_status < len(r) else None
        if status_raw and status_raw.lower() in EXCLUDE:
            continue
        qid = _s(r[c_id]) if c_id is not None and c_id < len(r) else None
        if not qid:
            continue
        rows.append(_row(
            f"PJM-{qid}",
            _s(r[c_name]) if c_name is not None and c_name < len(r) else None,
            "PJM",
            _s(r[c_state]) if c_state is not None and c_state < len(r) else None,
            _s(r[c_county]) if c_county is not None and c_county < len(r) else None,
            _s(r[c_fuel]) if c_fuel is not None and c_fuel < len(r) else None,
            _mw(r[c_mw]) if c_mw is not None and c_mw < len(r) else None,
            (status_raw or "active"),
            _date(r[c_date]) if c_date is not None and c_date < len(r) else None,
            None,
            src,
        ))
    return rows


# ══════════════════════════════════════════════════════════════════════
# SPP — active GI requests CSV (banner row 0, header row 1, data 2+)
# ══════════════════════════════════════════════════════════════════════
def parse_spp():
    url = "https://opsportal.spp.org/Studies/GenerateActiveCsv"
    src = "SPP GI active requests (live CSV)"
    body, st = _fetch(url, timeout=120)
    if st != 200 or not body:
        raise RuntimeError(f"fetch http_{st}")
    rows_csv = list(csv.reader(io.StringIO(body)))
    if len(rows_csv) < 3:
        raise RuntimeError("too few rows")
    idx = _hdr_index(rows_csv[1])  # row 0 is the 'Last Updated On' banner
    c_gin = _find(idx, "generation interconnection number")
    c_town = _find(idx, "nearest town or county")
    c_state = _find(idx, "state")
    c_mw = _find(idx, "max summer mw") or _find(idx, "capacity")
    c_fuel = _find(idx, "fuel type") or _find(idx, "generation type")
    c_status = _find(idx, "status")
    c_cess = _find(idx, "cessation date")
    c_poi = _find(idx, "substation or line")
    c_date = _find(idx, "request received")
    rows = []
    for r in rows_csv[2:]:
        if c_cess is not None and c_cess < len(r) and _s(r[c_cess]):
            continue  # cessation date present → withdrawn
        qid = _s(r[c_gin]) if c_gin is not None and c_gin < len(r) else None
        if not qid:
            continue
        rows.append(_row(
            f"SPP-{qid}",
            None,  # SPP active feed has no public project name
            "SPP",
            _s(r[c_state]) if c_state is not None and c_state < len(r) else None,
            _s(r[c_town]) if c_town is not None and c_town < len(r) else None,
            _s(r[c_fuel]) if c_fuel is not None and c_fuel < len(r) else None,
            _mw(r[c_mw]) if c_mw is not None and c_mw < len(r) else None,
            _s(r[c_status]) if c_status is not None and c_status < len(r) else "active",
            _date(r[c_date]) if c_date is not None and c_date < len(r) else None,
            _s(r[c_poi]) if c_poi is not None and c_poi < len(r) else None,
            src,
        ))
    return rows


# ══════════════════════════════════════════════════════════════════════
# ERCOT — GIS Report Large + Small Gen sheets (header row = the one w/ 'INR')
# All ERCOT projects are in Texas (state=TX).
# ══════════════════════════════════════════════════════════════════════
def parse_ercot():
    src = "ERCOT GIS Report (live XLSX)"
    LISTING = "https://www.ercot.com/misapp/GetReports.do?reportTypeId=15933"
    DL = "https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId="
    html, st = _fetch(LISTING, timeout=25)
    if st != 200 or len(html) < 2000:
        raise RuntimeError(f"listing http_{st}")
    listrows = re.findall(
        r'RPT\.00015933\.\d+\.(\d{8})\.\d+\.(GIS_Report_[A-Za-z0-9]+\.xlsx).*?doclookupId=(\d+)',
        html, re.S)
    if not listrows:
        raise RuntimeError("no GIS xlsx in listing")
    _pub, fname, docid = listrows[0]
    b, st = _fetch(DL + docid, timeout=150, return_bytes=True)
    if st != 200 or len(b) < 50000:
        raise RuntimeError(f"download http_{st} small={len(b)}")
    wb = _wb(b)
    rows = []
    for sheet in ("Project Details - Large Gen", "Project Details - Small Gen"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # find header row (contains 'INR')
        hdr_idx = None
        idx = None
        for ri, r in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
            cells = [str(c).strip() if c is not None else "" for c in r]
            if "INR" in cells:
                hdr_idx = ri
                idx = _hdr_index(cells)
                break
        if hdr_idx is None:
            continue
        c_inr = _find(idx, "inr")
        c_name = _find(idx, "project name")
        c_poi = _find(idx, "poi location")
        c_county = _find(idx, "county")
        c_cod = _find(idx, "projected cod")
        c_fuel = _find(idx, "fuel")
        c_tech = _find(idx, "technology")
        c_mw = _find(idx, "capacity (mw)") or _find(idx, "capacity")
        for r in ws.iter_rows(min_row=hdr_idx + 1, values_only=True):
            inr = _s(r[c_inr]) if c_inr is not None and c_inr < len(r) else None
            if not inr:
                continue
            fuel = (_s(r[c_fuel]) if c_fuel is not None and c_fuel < len(r) else None) \
                or (_s(r[c_tech]) if c_tech is not None and c_tech < len(r) else None)
            rows.append(_row(
                f"ERCOT-{inr}",
                _s(r[c_name]) if c_name is not None and c_name < len(r) else None,
                "ERCOT", "TX",
                _s(r[c_county]) if c_county is not None and c_county < len(r) else None,
                fuel,
                _mw(r[c_mw]) if c_mw is not None and c_mw < len(r) else None,
                "active",
                _date(r[c_cod]) if c_cod is not None and c_cod < len(r) else None,
                _s(r[c_poi]) if c_poi is not None and c_poi < len(r) else None,
                src,
            ))
    return rows


# ══════════════════════════════════════════════════════════════════════
# ISO-NE — IRTT Public Queue HTML table (cookie-jar opener required)
# ══════════════════════════════════════════════════════════════════════
def parse_iso_ne():
    import http.cookiejar
    from lxml import html as LH
    url = "https://irtt.iso-ne.com/reports/external"
    src = "ISO-NE IRTT Public Queue (live HTML)"
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "text/html"})
    with opener.open(req, timeout=90) as resp:
        body = resp.read()
        st = resp.status
    if st != 200 or len(body) < 50000:
        raise RuntimeError(f"fetch http_{st} small={len(body)}")
    doc = LH.fromstring(body)
    tbls = doc.xpath('//table[@id="publicqueue"]')
    if not tbls:
        raise RuntimeError("no publicqueue table")
    tbl = tbls[0]
    heads = [(th.text_content() or "").strip() for th in tbl.xpath('.//thead//th')]
    if not heads:
        heads = [(c.text_content() or "").strip()
                 for c in tbl.xpath('.//tr[1]/th | .//tr[1]/td')]
    idx = _hdr_index(heads)
    c_qp = _find(idx, "qp")
    c_status = _find(idx, "status")
    c_name = _find(idx, "alternative name")
    c_fuel = _find(idx, "fuel type")
    c_summer = _find(idx, "summer mw")
    c_net = _find(idx, "net mw")
    c_county = _find(idx, "county")
    c_st = _find(idx, "st")
    c_poi = _find(idx, "poi")
    c_date = _find(idx, "op date")
    rows = []
    STFUEL = {"WND": "Wind", "SUN": "Solar", "NG": "Natural Gas", "BAT": "Battery",
              "ES": "Energy Storage", "WAT": "Hydro", "NUC": "Nuclear"}
    for tr in tbl.xpath('.//tbody//tr'):
        tds = tr.xpath('./td')
        vals = [(td.text_content() or "").strip() for td in tds]
        if not vals:
            continue
        status = vals[c_status] if c_status is not None and c_status < len(vals) else ""
        if status.upper() != "A":  # active only
            continue
        qp = vals[c_qp] if c_qp is not None and c_qp < len(vals) else None
        if not _s(qp):
            continue
        mw = None
        for ci in (c_summer, c_net):
            if ci is not None and ci < len(vals):
                mw = _mw(vals[ci])
                if mw:
                    break
        fuel_raw = vals[c_fuel] if c_fuel is not None and c_fuel < len(vals) else None
        fuel = STFUEL.get((fuel_raw or "").upper(), fuel_raw)
        rows.append(_row(
            f"ISO-NE-{_s(qp)}",
            vals[c_name] if c_name is not None and c_name < len(vals) else None,
            "ISO-NE",
            vals[c_st] if c_st is not None and c_st < len(vals) else None,
            vals[c_county] if c_county is not None and c_county < len(vals) else None,
            fuel,
            mw,
            "active",
            _date(vals[c_date]) if c_date is not None and c_date < len(vals) else None,
            vals[c_poi] if c_poi is not None and c_poi < len(vals) else None,
            src,
        ))
    return rows


PARSERS = {
    "NYISO":  parse_nyiso,
    "CAISO":  parse_caiso,
    "MISO":   parse_miso,
    "PJM":    parse_pjm,
    "SPP":    parse_spp,
    "ERCOT":  parse_ercot,
    "ISO-NE": parse_iso_ne,
}


def _dedupe(rows):
    """Last write wins on (iso, queue_id), so the batch UPSERT never double-hits a key."""
    seen = {}
    for r in rows:
        seen[(r[2], r[0])] = r  # (iso, queue_id)
    return list(seen.values())


def ensure_schema(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS interconnect_queue_iso_qid_uidx
            ON interconnect_queue (iso, queue_id)
        """)
    conn.commit()


UPSERT_SQL = """
    INSERT INTO interconnect_queue
        (queue_id, project_name, iso, state, county, fuel_type,
         capacity_mw, queue_status, queue_date, poi_name, lat, lng,
         source, loaded_at)
    VALUES %s
    ON CONFLICT (iso, queue_id) DO UPDATE SET
        project_name = EXCLUDED.project_name,
        state        = EXCLUDED.state,
        county       = EXCLUDED.county,
        fuel_type    = EXCLUDED.fuel_type,
        capacity_mw  = EXCLUDED.capacity_mw,
        queue_status = EXCLUDED.queue_status,
        queue_date   = EXCLUDED.queue_date,
        poi_name     = EXCLUDED.poi_name,
        source       = EXCLUDED.source,
        loaded_at    = EXCLUDED.loaded_at
"""


def upsert(conn, rows):
    from psycopg2.extras import execute_values   # lazy: only the DB-write path needs psycopg2
    now = datetime.datetime.now(datetime.timezone.utc)
    payload = [r + (now,) for r in rows]  # append loaded_at
    with conn.cursor() as cur:
        execute_values(cur, UPSERT_SQL, payload,
                       template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                       page_size=1000)
    conn.commit()


def main(argv):
    args = argv[1:]
    emit_json = "--emit-json" in args
    targets = [a.upper() for a in args if not a.startswith("-")] or list(PARSERS.keys())

    if emit_json:
        # Parse-only: print {ISO: [rows...]} as JSON to stdout, NO DB. Used by the
        # GH-runner-fetch->POST fallback: the runner (open egress, no DATABASE_URL)
        # fetches+parses a Railway-blocked ISO, then POSTs these rows to
        # /api/v1/iso-queue/ingest-projects for upsert. queue_date -> ISO string.
        out = {}
        for iso in targets:
            fn = PARSERS.get(iso)
            if not fn:
                continue
            try:
                out[iso] = [list(r) for r in _dedupe(fn()) if r[0]]
            except Exception as e:
                out[iso] = {"error": f"{type(e).__name__}: {str(e)[:160]}"}
        print(json.dumps(out, default=str))
        return

    if not DATABASE_URL:
        raise SystemExit("Set DATABASE_URL (run via: railway run python3 ...)")
    import psycopg2
    conn = psycopg2.connect(DATABASE_URL)
    ensure_schema(conn)

    summary = {}
    for iso in targets:
        fn = PARSERS.get(iso)
        if not fn:
            print(f"[{iso}] SKIP — unknown ISO")
            summary[iso] = ("skipped", 0, "unknown ISO")
            continue
        try:
            rows = _dedupe(fn())
            rows = [r for r in rows if r[0]]  # require queue_id
            if not rows:
                print(f"[{iso}] parsed 0 rows — SKIP upsert")
                summary[iso] = ("empty", 0, "0 rows parsed")
                continue
            upsert(conn, rows)
            with_mw = sum(1 for r in rows if r[6] is not None)
            print(f"[{iso}] upserted {len(rows)} rows ({with_mw} with MW)")
            summary[iso] = ("ok", len(rows), f"{with_mw} with MW")
        except Exception as e:
            print(f"[{iso}] FAILED — {type(e).__name__}: {str(e)[:160]}")
            summary[iso] = ("failed", 0, f"{type(e).__name__}: {str(e)[:120]}")

    # ── post-load verification (the real, confirming count) ──
    print("\n" + "=" * 60)
    print("POST-LOAD COUNT(*) BY ISO (interconnect_queue):")
    with conn.cursor() as cur:
        cur.execute("SELECT iso, COUNT(*) FROM interconnect_queue GROUP BY iso ORDER BY COUNT(*) DESC")
        per_iso = cur.fetchall()
        for iso, n in per_iso:
            print(f"   {str(iso):8s}  {n:6d}")
        cur.execute("SELECT COUNT(*) FROM interconnect_queue")
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM interconnect_queue WHERE capacity_mw IS NOT NULL")
        with_mw = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM interconnect_queue WHERE project_name IS NOT NULL")
        with_name = cur.fetchone()[0]
    print(f"   {'TOTAL':8s}  {total:6d}   ({with_mw} with capacity_mw, {with_name} with project_name)")
    print("=" * 60)
    print("\nPER-ISO RUN RESULT:")
    for iso, (status, n, note) in summary.items():
        print(f"   {iso:8s} {status:8s} {note}")
    conn.close()


if __name__ == "__main__":
    main(sys.argv)
