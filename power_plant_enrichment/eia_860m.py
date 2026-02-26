"""
EIA-860M Monthly Generator Report + EIA-860 Annual - Lat/Lng Extraction
=========================================================================
Downloads EIA data and extracts exact coordinates for US power plants.

Two modes:
  1. EIA-860M monthly XLSX (direct download, not ZIP)
     URL: https://www.eia.gov/electricity/data/eia860m/xls/{month}_generator{year}.xlsx
  2. EIA-860 annual ZIP (contains Plant file with lat/lng)
     URL: https://www.eia.gov/electricity/data/eia860/xls/eia860{year}.zip

The annual EIA-860 Plant file is the authoritative source for coordinates.
The monthly 860M may or may not include lat/lng depending on the release.
"""

import csv
import io
import logging
import os
import tempfile
import zipfile
from datetime import datetime
from typing import Optional

import requests

logger = logging.getLogger(__name__)

EIA_860M_BASE_URL = "https://www.eia.gov/electricity/data/eia860m/xls"
EIA_860_ANNUAL_BASE_URL = "https://www.eia.gov/electricity/data/eia860/xls"

COLUMN_ALIASES = {
    "plant_id": ["Plant ID", "Plant Id", "Entity ID", "Plant Code", "plant_id"],
    "plant_name": ["Plant Name", "plant_name", "Station Name"],
    "state": ["Plant State", "State", "plant_state"],
    "latitude": ["Latitude", "latitude", "Lat"],
    "longitude": ["Longitude", "longitude", "Lon", "Long"],
    "capacity_mw": ["Nameplate Capacity (MW)", "Net Summer Capacity (MW)",
                     "nameplate_capacity_mw", "Capacity (MW)"],
    "technology": ["Technology", "Prime Mover", "technology"],
    "energy_source": ["Energy Source 1", "Energy Source Code",
                       "energy_source_1", "Fuel Type"],
    "status": ["Status", "Operating Status", "status", "Plant Status"],
    "county": ["County", "county"],
    "balancing_authority": ["Balancing Authority Code", "BA Code",
                             "balancing_authority_code"],
}


class EIA860MIngester:
    def __init__(self, cache_dir=None, api_key=None):
        self.cache_dir = cache_dir or tempfile.mkdtemp(prefix="eia860_")
        self.api_key = api_key or os.environ.get("EIA_API_KEY")
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "DCHub-DataEnrichment/1.0 (dchub.cloud)"
        })

    def _resolve_column(self, headers, field_name):
        aliases = COLUMN_ALIASES.get(field_name, [field_name])
        for alias in aliases:
            for i, h in enumerate(headers):
                if h and str(h).strip().lower() == alias.strip().lower():
                    return i
        return None

    def _build_column_map(self, headers):
        col_map = {}
        for field in COLUMN_ALIASES:
            col_map[field] = self._resolve_column(headers, field)
        return col_map

    def download_monthly_860m(self, year=None, month=None):
        """Download EIA-860M as direct XLSX (not ZIP)."""
        now = datetime.now()
        year = year or now.year
        month = month or max(1, now.month - 2)

        month_names = [
            "", "january", "february", "march", "april", "may", "june",
            "july", "august", "september", "october", "november", "december"
        ]

        # Build list of URLs to try (current month, then work backwards)
        url_patterns = []
        for m in range(month, max(0, month - 4), -1):
            if m < 1:
                break
            mn = month_names[m]
            url_patterns.append(f"{EIA_860M_BASE_URL}/{mn}_generator{year}.xlsx")
        # Also try prior year December
        url_patterns.append(f"{EIA_860M_BASE_URL}/december_generator{year - 1}.xlsx")

        xlsx_path = os.path.join(self.cache_dir, f"eia860m_{year}_{month:02d}.xlsx")

        for url in url_patterns:
            try:
                logger.info(f"Trying: {url}")
                resp = self.session.get(url, timeout=60)
                if resp.status_code == 200 and len(resp.content) > 50000:
                    # XLSX files start with PK (ZIP signature)
                    if resp.content[:2] == b"PK":
                        with open(xlsx_path, "wb") as f:
                            f.write(resp.content)
                        logger.info(f"Downloaded EIA-860M ({len(resp.content):,} bytes)")
                        return xlsx_path
                    else:
                        logger.debug(f"Not an XLSX file from {url}")
            except requests.RequestException as e:
                logger.debug(f"Failed: {url} -> {e}")
        
        raise FileNotFoundError(
            "Could not download EIA-860M. Check "
            "https://www.eia.gov/electricity/data/eia860m/ for available dates."
        )

    def download_annual_860(self, year=None):
        """Download annual EIA-860 ZIP and extract the Plant file (has lat/lng)."""
        now = datetime.now()
        year = year or now.year - 1

        urls = [
            f"{EIA_860_ANNUAL_BASE_URL}/eia860{year}.zip",
            f"{EIA_860_ANNUAL_BASE_URL}/eia860{year - 1}.zip",
        ]

        zip_path = os.path.join(self.cache_dir, f"eia860_{year}.zip")

        for url in urls:
            try:
                logger.info(f"Trying: {url}")
                resp = self.session.get(url, timeout=120)
                if resp.status_code == 200 and len(resp.content) > 100000:
                    with open(zip_path, "wb") as f:
                        f.write(resp.content)
                    logger.info(f"Downloaded EIA-860 annual ({len(resp.content):,} bytes)")
                    break
            except requests.RequestException as e:
                logger.debug(f"Failed: {url} -> {e}")
        else:
            raise FileNotFoundError(f"Could not download EIA-860 annual for {year}")

        # Extract the Plant file (best source for lat/lng)
        with zipfile.ZipFile(zip_path, "r") as zf:
            plant_files = [n for n in zf.namelist()
                           if "plant" in n.lower() and n.endswith(".xlsx")]
            gen_files = [n for n in zf.namelist()
                         if "generator" in n.lower() and n.endswith(".xlsx")]
            
            target = (plant_files or gen_files or
                      [n for n in zf.namelist() if n.endswith(".xlsx")])[0]
            extract_path = zf.extract(target, self.cache_dir)
            logger.info(f"Extracted: {target}")
            return extract_path

    def parse_plants_from_xlsx(self, xlsx_path):
        """Parse XLSX and return plant records with coordinates."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required: add to requirements.txt")

        plants = {}
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Find best sheet
        target_sheet = None
        for name in wb.sheetnames:
            nl = name.lower()
            if "operat" in nl or "plant" in nl:
                target_sheet = name
                break
        target_sheet = target_sheet or wb.sheetnames[0]

        logger.info(f"Parsing sheet: '{target_sheet}'")
        ws = wb[target_sheet]

        # Find header row (scan first 10 rows)
        headers = None
        col_map = None
        header_row = 0

        for row in ws.iter_rows(max_row=10, values_only=True):
            header_row += 1
            row_strs = [str(c).strip() if c else "" for c in row]
            if any("plant" in cell.lower() for cell in row_strs[:15] if cell):
                headers = row_strs
                col_map = self._build_column_map(headers)
                if col_map.get("plant_id") is not None:
                    break

        if not col_map or col_map.get("plant_id") is None:
            wb.close()
            raise ValueError(f"Could not find Plant ID column in {xlsx_path}. "
                             f"Sheets: {wb.sheetnames}")

        lat_idx = col_map.get("latitude")
        lng_idx = col_map.get("longitude")
        id_idx = col_map["plant_id"]
        has_coords = lat_idx is not None and lng_idx is not None

        logger.info(f"Columns: plant_id={id_idx}, lat={lat_idx}, lng={lng_idx}")

        # Parse data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            try:
                if not row or len(row) <= id_idx:
                    continue

                plant_id = str(row[id_idx]).strip() if row[id_idx] else ""
                if not plant_id or plant_id in ("None", ""):
                    continue

                lat_f, lng_f = None, None
                if has_coords:
                    lat_f = _safe_float(row[lat_idx] if lat_idx < len(row) else None)
                    lng_f = _safe_float(row[lng_idx] if lng_idx < len(row) else None)
                    if lat_f is not None and not (-90 <= lat_f <= 90):
                        lat_f, lng_f = None, None
                    if lng_f is not None and not (-180 <= lng_f <= 180):
                        lat_f, lng_f = None, None

                def _get(field):
                    idx = col_map.get(field)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return None

                if plant_id not in plants:
                    plants[plant_id] = {
                        "eia_plant_id": plant_id,
                        "name": _get("plant_name"),
                        "state": _get("state"),
                        "county": _get("county"),
                        "latitude": lat_f,
                        "longitude": lng_f,
                        "capacity_mw": _safe_float(_get("capacity_mw")),
                        "technology": _get("technology"),
                        "energy_source": _get("energy_source"),
                        "status": _get("status"),
                        "balancing_authority": _get("balancing_authority"),
                        "source": "eia_860",
                        "updated_at": datetime.utcnow().isoformat(),
                    }
                else:
                    cap = _safe_float(_get("capacity_mw"))
                    if cap and plants[plant_id].get("capacity_mw"):
                        plants[plant_id]["capacity_mw"] += cap
                    elif cap:
                        plants[plant_id]["capacity_mw"] = cap
                    if lat_f and plants[plant_id].get("latitude") is None:
                        plants[plant_id]["latitude"] = lat_f
                        plants[plant_id]["longitude"] = lng_f

            except (ValueError, IndexError):
                continue

        wb.close()
        result = list(plants.values())
        with_coords = sum(1 for p in result if p.get("latitude") is not None)
        logger.info(f"Parsed {len(result)} plants ({with_coords} with coordinates)")
        return result

    def ingest(self, year=None, month=None, use_annual=True):
        """
        Full pipeline: download → parse → return plants.
        Defaults to annual EIA-860 (best for coordinates).
        Falls back to monthly 860M if annual fails.
        """
        # Try annual first (better lat/lng coverage)
        if use_annual:
            try:
                xlsx_path = self.download_annual_860(year=year)
                return self.parse_plants_from_xlsx(xlsx_path)
            except Exception as e:
                logger.warning(f"Annual EIA-860 failed ({e}), trying monthly 860M...")

        # Fall back to monthly
        xlsx_path = self.download_monthly_860m(year=year, month=month)
        return self.parse_plants_from_xlsx(xlsx_path)


def _safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
